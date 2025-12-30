"""
Picklist Summary Tab Helper
Generates summary statistics for picklist exports
"""
from typing import Dict, List, Tuple
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from excel_style_helper import ExcelStyleHelper
from datetime import datetime


class PicklistSummaryData:
    """Stores summary statistics for a single object"""
    
    def __init__(self, object_name: str, object_api: str):
        self.object_name = object_name
        self.object_api = object_api
        self.total_picklist_fields = 0  # ✅ NEW: Total count of all picklist fields
        self.standard_picklist_count = 0
        self.global_picklist_count = 0
        self.dependent_picklist_count = 0
        self.custom_picklist_count = 0
        self.active_values = 0
        self.inactive_values = 0
        self.object_type = "Custom" if object_api.endswith("__c") else "Standard"
    
    def to_row(self, serial_number: int) -> List:
        """Convert to Excel row format"""
        return [
            serial_number,
            self.object_name,
            self.object_api,
            self.total_picklist_fields,  # ✅ NEW: Total Picklist Fields column
            self.standard_picklist_count,
            self.global_picklist_count,
            self.dependent_picklist_count,
            self.custom_picklist_count,
            self.active_values,
            self.inactive_values,
            self.object_type
        ]


class PicklistSummaryHelper:
    """Helper class for creating picklist summary tabs"""
    
    # ✅ UPDATED: Summary tab headers with new column
    SUMMARY_HEADERS = [
        'SL',
        'Object Name',
        'Object API',
        'Total Picklist Fields',  # ✅ NEW COLUMN
        'Standard Picklist',
        'Global Picklist',
        'Dependent Picklist',
        'Custom Picklist',
        'Active Values',
        'Inactive Values',
        'Object Type'
    ]
    
    @staticmethod
    def create_summary_sheet(wb: Workbook, summary_data: List[PicklistSummaryData], 
                           total_stats: Dict) -> Worksheet:
        """
        Create summary sheet as first tab in workbook
        
        Args:
            wb: Workbook instance
            summary_data: List of PicklistSummaryData objects
            total_stats: Dictionary with overall statistics
            
        Returns:
            Created worksheet
        """
        # Create sheet at the beginning
        ws = wb.create_sheet("Summary", 0)
        
        headers = PicklistSummaryHelper.SUMMARY_HEADERS
        num_cols = len(headers)
        
        # ✅ Add Title Row (Row 1)
        ExcelStyleHelper.add_title_row(
            ws,
            title="Salesforce Picklist Export - Summary",
            num_cols=num_cols,
            row_num=1
        )
        
        # ✅ Add Info Row (Row 2)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
        info_cell = ws.cell(row=2, column=1)
        export_date = datetime.now().strftime('%Y-%m-%d %H:%M')
        
        total_objects = len(summary_data)
        total_picklist_fields = total_stats.get('total_picklist_fields', 0)
        total_values = total_stats.get('total_values', 0)
        
        info_cell.value = (
            f"Total Objects: {total_objects} | "
            f"Total Picklist Fields: {total_picklist_fields} | "
            f"Total Values: {total_values} | "
            f"Export Date: {export_date}"
        )
        
        # Apply info style
        info_style = ExcelStyleHelper.get_info_style()
        ExcelStyleHelper.apply_style_to_cell(info_cell, info_style)
        for col_num in range(2, num_cols + 1):
            cell = ws.cell(row=2, column=col_num)
            ExcelStyleHelper.apply_style_to_cell(cell, info_style)
        ws.row_dimensions[2].height = 20
        
        # ✅ Add Header Row (Row 3)
        ExcelStyleHelper.add_header_row(ws, headers, row_num=3)
        
        # ✅ Add Data Rows with alternating colors
        for idx, summary_obj in enumerate(summary_data, start=1):
            row_num = idx + 3  # Start from row 4
            row_data = summary_obj.to_row(idx)
            
            is_even_row = (row_num % 2 == 0)
            data_style = ExcelStyleHelper.get_data_style(is_even_row)
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.value = value
                ExcelStyleHelper.apply_style_to_cell(cell, data_style)
        
        # ✅ Add Totals Row (if data exists)
        if summary_data:
            totals_row_num = len(summary_data) + 4
            
            # ✅ Calculate totals (including new Total Picklist Fields)
            total_fields = sum(s.total_picklist_fields for s in summary_data)
            total_standard = sum(s.standard_picklist_count for s in summary_data)
            total_global = sum(s.global_picklist_count for s in summary_data)
            total_dependent = sum(s.dependent_picklist_count for s in summary_data)
            total_custom = sum(s.custom_picklist_count for s in summary_data)
            total_active = sum(s.active_values for s in summary_data)
            total_inactive = sum(s.inactive_values for s in summary_data)
            
            totals_data = [
                "",  # SL
                "TOTAL",  # Object Name
                "",  # Object API
                total_fields,  # ✅ NEW: Total Picklist Fields
                total_standard,
                total_global,
                total_dependent,
                total_custom,
                total_active,
                total_inactive,
                ""  # Object Type
            ]
            
            # Apply bold style to totals row
            from openpyxl.styles import Font, PatternFill
            totals_style = {
                'font': Font(name='Calibri', size=11, bold=True),
                'fill': PatternFill(
                    start_color="E0E0E0",
                    end_color="E0E0E0",
                    fill_type="solid"
                ),
                'alignment': ExcelStyleHelper.get_data_style()['alignment'],
                'border': ExcelStyleHelper.get_data_style()['border']
            }
            
            for col_idx, value in enumerate(totals_data, start=1):
                cell = ws.cell(row=totals_row_num, column=col_idx)
                cell.value = value
                ExcelStyleHelper.apply_style_to_cell(cell, totals_style)
        
        # ✅ Auto-adjust column widths
        ExcelStyleHelper.auto_adjust_column_widths(ws, headers, max_width=40)
        
        # ✅ Freeze header rows
        ExcelStyleHelper.freeze_header_rows(ws, num_rows=3)
        
        return ws
    
    @staticmethod
    def analyze_picklist_data(object_api: str, rows: List[List[str]], 
                            sf_client) -> PicklistSummaryData:
        """
        Analyze picklist data for an object and generate summary statistics
        
        Args:
            object_api: Object API name
            rows: List of data rows [Object, Field Label, Field API, Value Label, Value API, Status, IsGlobal?]
            sf_client: SalesforceClient instance for getting object label
            
        Returns:
            PicklistSummaryData object
        """
        # Get object label
        object_label = ExcelStyleHelper.get_object_label(sf_client, object_api)
        
        # Create summary data object
        summary = PicklistSummaryData(object_label, object_api)
        
        # Track unique fields
        seen_fields = set()
        field_info = {}  # field_api -> {is_global, is_dependent, is_standard, is_custom}
        
        for row in rows:
            if len(row) < 7:
                continue
            
            field_api = row[2]  # Field API column
            status = row[5]     # Status column
            is_global = row[6]  # IsGlobal? column
            
            # Count values
            if status == 'Active':
                summary.active_values += 1
            elif status == 'Inactive':
                summary.inactive_values += 1
            
            # Track unique fields
            if field_api not in seen_fields:
                seen_fields.add(field_api)
                
                # Determine field type
                is_custom = field_api.endswith('__c')
                is_standard = not is_custom
                is_global_field = (is_global == 'Yes')
                
                # Store field info for dependent picklist check
                field_info[field_api] = {
                    'is_global': is_global_field,
                    'is_standard': is_standard,
                    'is_custom': is_custom
                }
                
                # Count by type
                if is_global_field:
                    summary.global_picklist_count += 1
                
                if is_standard:
                    summary.standard_picklist_count += 1
                elif is_custom:
                    summary.custom_picklist_count += 1
        
        # ✅ NEW: Set total picklist fields count
        summary.total_picklist_fields = len(seen_fields)
        
        # ✅ Check for dependent picklists (requires Salesforce API call)
        summary.dependent_picklist_count = PicklistSummaryHelper._count_dependent_picklists(
            sf_client, 
            object_api, 
            list(seen_fields)
        )
        
        return summary
    
    @staticmethod
    def _count_dependent_picklists(sf_client, object_api: str, 
                                  field_apis: List[str]) -> int:
        """
        Count dependent picklist fields
        
        Args:
            sf_client: SalesforceClient instance
            object_api: Object API name
            field_apis: List of field API names to check
            
        Returns:
            Count of dependent picklist fields
        """
        try:
            obj_describe = getattr(sf_client.sf, object_api).describe()
            
            dependent_count = 0
            
            for field in obj_describe['fields']:
                if field['name'] in field_apis:
                    # Check if field is dependent
                    if field.get('dependentPicklist', False):
                        dependent_count += 1
            
            return dependent_count
            
        except Exception as e:
            # If error, return 0
            return 0