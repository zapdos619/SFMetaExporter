"""
Object metadata export functionality
"""
import csv
from typing import List, Dict, Tuple
from openpyxl import Workbook
import zipfile
import os
from datetime import datetime

from models import MetadataField
from salesforce_client import SalesforceClient
from field_usage_tracker import FieldUsageTracker
from excel_style_helper import ExcelStyleHelper
from metadata_summary_helper import MetadataSummaryHelper, MetadataSummaryData


class MetadataExporter:
    """Handles object metadata export from Salesforce"""

    # ✅ UPDATED: New headers with all columns
    METADATA_HEADERS = [
        'Object',
        'Field Label',
        'API Name',
        'Data Type',  # ✅ NEW
        'Length',  # ✅ NEW
        'Field Type',  # ✅ ENHANCED (more detailed)
        'Required',  # ✅ NEW
        'Picklist Values',  # ✅ NEW
        'Formula',
        'External ID',  # ✅ NEW (was "Extend ID")
        'Track History',  # ✅ NEW
        'Description',  # ✅ NEW
        'Help Text',
        'Attributes',
        'Field Usage'
    ]
    
    def __init__(self, sf_client: SalesforceClient):
        """Initialize with Salesforce client"""
        self.sf_client = sf_client
        self.sf = sf_client.sf
        self.usage_tracker = FieldUsageTracker(self.sf, sf_client.status_callback)

    
    def export_metadata(self, object_names: List[str], output_path: str) -> Tuple[str, Dict]:
        """
        Export metadata for specified objects
        
        ⚠️ LEGACY METHOD: This uses CSV format.
        New code should use export_metadata_excel() instead.
        
        Args:
            object_names: List of object API names
            output_path: Output CSV file path
            
        Returns:
            Tuple of (output_path, statistics_dict)
        """
        self._log_status("=== Starting Metadata Export ===")
        self._log_status(f"Total objects to process: {len(object_names)}")
        
        stats = {
            'total_objects': len(object_names),
            'successful_objects': 0,
            'failed_objects': 0,
            'total_fields': 0,
            'failed_object_details': []
        }
        
        all_metadata_fields: List[MetadataField] = []
        
        for i, obj_name in enumerate(object_names, 1):
            self._log_status(f"[{i}/{len(object_names)}] Processing object: {obj_name}")
            try:
                fields = self._get_object_metadata(obj_name)
                all_metadata_fields.extend(fields)
                stats['successful_objects'] += 1
                stats['total_fields'] += len(fields)
                self._log_status(f"  ✅ Retrieved {len(fields)} fields")
            except Exception as e:
                error_msg = str(e)
                self._log_status(f"  ❌ ERROR: {error_msg}")
                stats['failed_objects'] += 1
                stats['failed_object_details'].append({'name': obj_name, 'reason': error_msg})
            self._log_status("")
        
        self._log_status("=== Creating CSV File ===")
        final_output_path = self._create_csv_file(all_metadata_fields, output_path)
        return final_output_path, stats
    
  
    def export_metadata_excel(self, object_names: List[str], output_path: str, 
                            export_mode: str = "single_tab") -> Tuple[str, Dict]:
        """
        Export metadata for specified objects to Excel with styling
        
        ✅ UPDATED: Now includes Summary Tab
        
        Args:
            object_names: List of object API names
            output_path: Path for output file(s)
            export_mode: "single_tab", "multi_tab", or "individual_files"
            
        Returns:
            Tuple of (output_path, statistics_dict)
        """
        self._log_status("=== Starting Metadata Export (Excel Format) ===")
        self._log_status(f"Export Mode: {export_mode}")
        self._log_status(f"Total objects to process: {len(object_names)}")
        
        stats = {
            'total_objects': len(object_names),
            'successful_objects': 0,
            'failed_objects': 0,
            'total_fields': 0,
            'failed_object_details': [],
            'export_mode': export_mode
        }
        
        # Dispatch to appropriate export method based on mode
        if export_mode == "single_tab":
            return self._export_single_tab_with_summary(object_names, output_path, stats)
        elif export_mode == "multi_tab":
            return self._export_multi_tab_with_summary(object_names, output_path, stats)
        elif export_mode == "individual_files":
            return self._export_individual_files_with_summary(object_names, output_path, stats)
        else:
            raise ValueError(f"Invalid export mode: {export_mode}")


    def _export_single_tab_with_summary(self, object_names: List[str], 
                                       output_path: str, stats: Dict) -> Tuple[str, Dict]:
        """
        ✅ UPDATED: Single tab export with Summary tab
        """
        self._log_status("📄 Single Tab Mode: All objects in one sheet")
        
        # Collect all metadata fields AND summary data
        all_metadata_fields = []
        summary_data_list = []
        
        for i, obj_name in enumerate(sorted(object_names), 1):
            self._log_status(f"[{i}/{len(object_names)}] Processing object: {obj_name}")
            
            try:
                fields = self._get_object_metadata(obj_name)
                all_metadata_fields.extend(fields)
                stats['successful_objects'] += 1
                stats['total_fields'] += len(fields)
                
                # ✅ Generate summary data for this object
                summary_obj = MetadataSummaryHelper.analyze_metadata(
                    obj_name,
                    fields,
                    self.sf_client
                )
                summary_data_list.append(summary_obj)
                
                self._log_status(f"  ✅ Retrieved {len(fields)} fields")
                
            except Exception as e:
                error_msg = str(e)
                self._log_status(f"  ❌ ERROR: {error_msg}")
                stats['failed_objects'] += 1
                stats['failed_object_details'].append({
                    'name': obj_name, 
                    'reason': error_msg
                })
            
            self._log_status("")
        
        # Create Excel file with Summary tab + Data tab
        self._log_status("=== Creating Excel File with Summary ===")
        final_output_path = self._create_single_tab_excel_with_summary(
            all_metadata_fields,
            summary_data_list,
            output_path,
            stats
        )
        
        return final_output_path, stats
    
    def _create_single_tab_excel_with_summary(
        self, 
        metadata_fields: List[MetadataField],
        summary_data: List[MetadataSummaryData],
        output_path: str, 
        stats: Dict
    ) -> str:
        """
        ✅ NEW: Create Excel with Summary tab (first) + Data tab (second)
        """
        wb = Workbook()
        
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # ✅ Create Summary Tab (first)
        if summary_data:
            self._log_status("📊 Creating Summary tab...")
            MetadataSummaryHelper.create_summary_sheet(wb, summary_data, stats)
        
        # ✅ Create Data Tab (second)
        self._log_status("📋 Creating Metadata tab...")
        ws_data = wb.create_sheet("Metadata")
        
        headers = self.METADATA_HEADERS
        num_cols = len(headers)
        
        # Add Title Row
        ExcelStyleHelper.add_title_row(
            ws_data,
            title="Salesforce Metadata Export",
            num_cols=num_cols,
            row_num=1
        )
        
        # Add Info Row
        total_objects = stats['successful_objects']
        total_fields = len(metadata_fields)
        
        ws_data.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
        info_cell = ws_data.cell(row=2, column=1)
        export_date = datetime.now().strftime('%Y-%m-%d %H:%M')
        info_cell.value = (
            f"Objects: {total_objects} | "
            f"Total Fields: {total_fields} | "
            f"Export Date: {export_date}"
        )
        
        info_style = ExcelStyleHelper.get_info_style()
        ExcelStyleHelper.apply_style_to_cell(info_cell, info_style)
        for col_num in range(2, num_cols + 1):
            cell = ws_data.cell(row=2, column=col_num)
            ExcelStyleHelper.apply_style_to_cell(cell, info_style)
        ws_data.row_dimensions[2].height = 20
        
        # Add Header Row
        ExcelStyleHelper.add_header_row(ws_data, headers, row_num=3)
        
        # Add Data Rows with alternating colors
        for row_idx, field in enumerate(metadata_fields, start=4):
            row_data = field.to_row()
            is_even_row = (row_idx % 2 == 0)
            data_style = ExcelStyleHelper.get_data_style(is_even_row)
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_data.cell(row=row_idx, column=col_idx)
                cell.value = value
                ExcelStyleHelper.apply_style_to_cell(cell, data_style)
        
        # Auto-adjust column widths
        ExcelStyleHelper.auto_adjust_column_widths(ws_data, headers)
        
        # Freeze header rows
        ExcelStyleHelper.freeze_header_rows(ws_data, num_rows=3)
        
        # Save workbook
        wb.save(output_path)
        
        self._log_status(f"✅ Excel file created: {output_path}")
        self._log_status(f"✅ Total sheets: {len(wb.worksheets)} (Summary + Metadata)")
        
        return output_path
    
    def _export_multi_tab_with_summary(self, object_names: List[str], 
                                      output_path: str, stats: Dict) -> Tuple[str, Dict]:
        """
        ✅ UPDATED: Multiple tabs export with Summary tab
        """
        self._log_status("📑 Multiple Tabs Mode: One sheet per object + Summary")
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # Sort objects alphabetically
        sorted_objects = sorted(object_names)
        
        # Store summary data
        summary_data_list = []
        
        # Process each object and create a sheet
        for i, obj_name in enumerate(sorted_objects, 1):
            self._log_status(f"[{i}/{len(sorted_objects)}] Processing object: {obj_name}")
            
            try:
                fields = self._get_object_metadata(obj_name)
                stats['successful_objects'] += 1
                stats['total_fields'] += len(fields)
                
                # Get object label for tab name
                object_label = ExcelStyleHelper.get_object_label(self.sf_client, obj_name)
                
                # Create sheet for this object
                self._create_sheet_for_object(
                    wb,
                    obj_name,
                    object_label,
                    fields
                )
                
                # ✅ Generate summary data for this object
                summary_obj = MetadataSummaryHelper.analyze_metadata(
                    obj_name,
                    fields,
                    self.sf_client
                )
                summary_data_list.append(summary_obj)
                
                self._log_status(f"  ✅ Retrieved {len(fields)} fields")
                
            except Exception as e:
                error_msg = str(e)
                self._log_status(f"  ❌ ERROR: {error_msg}")
                stats['failed_objects'] += 1
                stats['failed_object_details'].append({
                    'name': obj_name,
                    'reason': error_msg
                })
            
            self._log_status("")
        
        # ✅ Create Summary Tab (first tab)
        if summary_data_list:
            self._log_status("📊 Creating Summary tab...")
            MetadataSummaryHelper.create_summary_sheet(wb, summary_data_list, stats)
        else:
            # No data - create placeholder
            ws = wb.create_sheet("No Data", 0)
            ws['A1'] = "No metadata found for selected objects"
            self._log_status("⚠️ No data exported - created placeholder sheet")
        
        # Save workbook
        wb.save(output_path)
        
        self._log_status(f"✅ Excel file created: {output_path}")
        self._log_status(f"✅ Total sheets: {len(wb.worksheets)} (Summary + {len(summary_data_list)} objects)")
        
        return output_path, stats    


    def _export_individual_files_with_summary(self, object_names: List[str], 
                                             output_path: str, stats: Dict) -> Tuple[str, Dict]:
        """
        ✅ UPDATED: Individual files export with separate Summary file
        """
        self._log_status("📦 Individual Files Mode: Separate .xlsx per object + Summary file")
        
        # Sort objects alphabetically
        sorted_objects = sorted(object_names)
        
        # Determine output strategy
        export_type = "Metadata"
        
        # Get base directory
        output_dir = os.path.dirname(output_path)
        
        # List to store created file paths
        created_files = []
        summary_data_list = []
        
        # Process each object
        for i, obj_name in enumerate(sorted_objects, 1):
            self._log_status(f"[{i}/{len(sorted_objects)}] Processing object: {obj_name}")
            
            try:
                fields = self._get_object_metadata(obj_name)
                stats['successful_objects'] += 1
                stats['total_fields'] += len(fields)
                
                # Get object label
                object_label = ExcelStyleHelper.get_object_label(self.sf_client, obj_name)
                
                # Create individual Excel file
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{obj_name}_{export_type}_{timestamp}.xlsx"
                file_path = os.path.join(output_dir, filename)
                
                self._create_individual_excel_file(
                    file_path,
                    obj_name,
                    object_label,
                    fields
                )
                
                created_files.append(file_path)
                
                # ✅ Generate summary data
                summary_obj = MetadataSummaryHelper.analyze_metadata(
                    obj_name,
                    fields,
                    self.sf_client
                )
                summary_data_list.append(summary_obj)
                
                self._log_status(
                    f"  ✅ Created: {filename} | Fields: {len(fields)}"
                )
                
            except Exception as e:
                error_msg = str(e)
                self._log_status(f"  ❌ ERROR: {error_msg}")
                stats['failed_objects'] += 1
                stats['failed_object_details'].append({
                    'name': obj_name,
                    'reason': error_msg
                })
            
            self._log_status("")
        
        # ✅ Create Summary File
        if summary_data_list:
            self._log_status("📊 Creating Summary file...")
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            summary_filename = f"Summary_{export_type}_{timestamp}.xlsx"
            summary_path = os.path.join(output_dir, summary_filename)
            
            wb_summary = Workbook()
            default_sheet = wb_summary.active
            wb_summary.remove(default_sheet)
            
            MetadataSummaryHelper.create_summary_sheet(wb_summary, summary_data_list, stats)
            wb_summary.save(summary_path)
            
            created_files.insert(0, summary_path)  # Add summary at beginning
            self._log_status(f"✅ Summary file created: {summary_filename}")
        
        # Decide: Single file or ZIP?
        if len(created_files) == 0:
            self._log_status("⚠️ No files created - no data to export")
            raise Exception("No metadata found for any selected objects")
            
        elif len(created_files) == 1:
            # Single file (just summary)
            final_path = created_files[0]
            self._log_status(f"✅ Single file created: {final_path}")
            return final_path, stats
            
        else:
            # Multiple files - create ZIP
            self._log_status(f"📦 Creating ZIP archive for {len(created_files)} files...")
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            zip_filename = f"Salesforce_{export_type}_Export_{timestamp}.zip"
            zip_path = os.path.join(output_dir, zip_filename)
            
            final_path = self._create_zip_archive(created_files, zip_path)
            
            return final_path, stats
    
    def _create_zip_archive(self, file_paths: List[str], zip_path: str) -> str:
        """
        Create ZIP archive containing multiple Excel files
        
        Args:
            file_paths: List of file paths to include in ZIP
            zip_path: Output ZIP file path
            
        Returns:
            Path to created ZIP file
        """
        try:
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for file_path in file_paths:
                    arcname = os.path.basename(file_path)
                    zipf.write(file_path, arcname=arcname)
                    self._log_status(f"  📄 Added to ZIP: {arcname}")
            
            # Delete individual files after zipping
            self._log_status("🧹 Cleaning up individual files...")
            for file_path in file_paths:
                try:
                    os.remove(file_path)
                except Exception as e:
                    self._log_status(f"  ⚠️ Could not delete {file_path}: {e}")
            
            self._log_status(f"✅ ZIP archive created: {zip_path}")
            self._log_status(f"✅ Total files in ZIP: {len(file_paths)}")
            
            return zip_path
            
        except Exception as e:
            self._log_status(f"❌ Error creating ZIP: {str(e)}")
            raise
    
    
    def _export_single_tab(self, object_names: List[str], output_path: str, 
                        stats: Dict) -> Tuple[str, Dict]:
        """
        Export all objects to a single Excel sheet with styling
        
        Args:
            object_names: List of object API names
            output_path: Output Excel file path
            stats: Statistics dictionary
            
        Returns:
            Tuple of (output_path, stats)
        """
        self._log_status("📄 Single Tab Mode: All objects in one sheet")
        
        # Collect all metadata fields
        all_metadata_fields: List[MetadataField] = []
        
        for i, obj_name in enumerate(sorted(object_names), 1):  # ✅ Sort alphabetically
            self._log_status(f"[{i}/{len(object_names)}] Processing object: {obj_name}")
            try:
                fields = self._get_object_metadata(obj_name)
                all_metadata_fields.extend(fields)
                stats['successful_objects'] += 1
                stats['total_fields'] += len(fields)
                self._log_status(f"  ✅ Retrieved {len(fields)} fields")
            except Exception as e:
                error_msg = str(e)
                self._log_status(f"  ❌ ERROR: {error_msg}")
                stats['failed_objects'] += 1
                stats['failed_object_details'].append({'name': obj_name, 'reason': error_msg})
            self._log_status("")
        
        # Create Excel file with styling
        self._log_status("=== Creating Excel File ===")
        final_output_path = self._create_single_tab_excel(all_metadata_fields, output_path, stats)
        
        return final_output_path, stats    
    
    
    def _create_single_tab_excel(self, metadata_fields: List[MetadataField], 
                                output_path: str, stats: Dict) -> str:
        """
        Create Excel file with single tab and styling
        
        ✅ UPDATED: Use new METADATA_HEADERS
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Metadata Export"
        
        # ✅ Use class constant for headers
        headers = self.METADATA_HEADERS
        num_cols = len(headers)
        
        # Add Title Row
        ExcelStyleHelper.add_title_row(
            ws,
            title="Salesforce Metadata Export",
            num_cols=num_cols,
            row_num=1
        )
        
        # Add Info Row
        total_objects = stats['successful_objects']
        total_fields = len(metadata_fields)
        
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
        info_cell = ws.cell(row=2, column=1)
        export_date = datetime.now().strftime('%Y-%m-%d %H:%M')
        info_cell.value = (
            f"Objects: {total_objects} | "
            f"Total Fields: {total_fields} | "
            f"Export Date: {export_date}"
        )
        
        # Apply info style
        info_style = ExcelStyleHelper.get_info_style()
        ExcelStyleHelper.apply_style_to_cell(info_cell, info_style)
        for col_num in range(2, num_cols + 1):
            cell = ws.cell(row=2, column=col_num)
            ExcelStyleHelper.apply_style_to_cell(cell, info_style)
        ws.row_dimensions[2].height = 20
        
        # Add Header Row
        ExcelStyleHelper.add_header_row(ws, headers, row_num=3)
        
        # Add Data Rows with alternating colors
        for row_idx, field in enumerate(metadata_fields, start=4):
            row_data = field.to_row()
            is_even_row = (row_idx % 2 == 0)
            data_style = ExcelStyleHelper.get_data_style(is_even_row)
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                ExcelStyleHelper.apply_style_to_cell(cell, data_style)
        
        # Auto-adjust column widths
        ExcelStyleHelper.auto_adjust_column_widths(ws, headers)
        
        # Freeze header rows
        ExcelStyleHelper.freeze_header_rows(ws, num_rows=3)
        
        # Save workbook
        wb.save(output_path)
        
        self._log_status(f"✅ Excel file created: {output_path}")
        self._log_status(f"✅ Total fields exported: {len(metadata_fields)}")
        
        return output_path
    
 
    def _export_multi_tab(self, object_names: List[str], output_path: str, 
                        stats: Dict) -> Tuple[str, Dict]:
        """
        Export each object to its own sheet in a single Excel file
        
        Args:
            object_names: List of object API names
            output_path: Output Excel file path
            stats: Statistics dictionary
            
        Returns:
            Tuple of (output_path, stats)
        """
        self._log_status("📑 Multiple Tabs Mode: One sheet per object")
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Sort objects alphabetically
        sorted_objects = sorted(object_names)
        
        # Process each object and create a sheet
        for i, obj_name in enumerate(sorted_objects, 1):
            self._log_status(f"[{i}/{len(sorted_objects)}] Processing object: {obj_name}")
            
            try:
                fields = self._get_object_metadata(obj_name)
                stats['successful_objects'] += 1
                stats['total_fields'] += len(fields)
                
                # ✅ Get object label for tab name
                object_label = ExcelStyleHelper.get_object_label(self.sf_client, obj_name)
                
                # ✅ Create sheet for this object
                self._create_sheet_for_object(
                    wb,
                    obj_name,
                    object_label,
                    fields
                )
                
                self._log_status(f"  ✅ Retrieved {len(fields)} fields")
                
            except Exception as e:
                error_msg = str(e)
                self._log_status(f"  ❌ ERROR: {error_msg}")
                stats['failed_objects'] += 1
                stats['failed_object_details'].append({'name': obj_name, 'reason': error_msg})
            
            self._log_status("")
        
        # Save workbook
        if len(wb.worksheets) == 0:
            # No sheets created - create a placeholder
            ws = wb.create_sheet("No Data")
            ws['A1'] = "No metadata found for selected objects"
            self._log_status("⚠️ No data exported - created placeholder sheet")
        
        wb.save(output_path)
        
        self._log_status(f"✅ Excel file created: {output_path}")
        self._log_status(f"✅ Total sheets: {len(wb.worksheets)}")
        
        return output_path, stats    
    
    
    def _create_sheet_for_object(self, wb: Workbook, object_api: str, 
                                object_label: str, fields: List[MetadataField]):
        """
        Create a single sheet for an object with styling
        
        ✅ UPDATED: Use new METADATA_HEADERS
        """
        # Sanitize sheet name
        sheet_name = ExcelStyleHelper.sanitize_sheet_name(object_label)
        
        # Create sheet
        ws = wb.create_sheet(sheet_name)
        
        # ✅ Use class constant for headers
        headers = self.METADATA_HEADERS
        num_cols = len(headers)
        
        # Add Title Row
        ExcelStyleHelper.add_title_row(
            ws,
            title=f"Salesforce Metadata Export - {object_label}",
            num_cols=num_cols,
            row_num=1
        )
        
        # Add Info Row with object details
        ExcelStyleHelper.add_info_row(
            ws,
            object_label=object_label,
            object_api=object_api,
            record_count=len(fields),
            num_cols=num_cols,
            row_num=2
        )
        
        # Add Header Row
        ExcelStyleHelper.add_header_row(ws, headers, row_num=3)
        
        # Add Data Rows with alternating colors
        for row_idx, field in enumerate(fields, start=4):
            row_data = field.to_row()
            is_even_row = (row_idx % 2 == 0)
            data_style = ExcelStyleHelper.get_data_style(is_even_row)
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                ExcelStyleHelper.apply_style_to_cell(cell, data_style)
        
        # Auto-adjust column widths
        ExcelStyleHelper.auto_adjust_column_widths(ws, headers)
        
        # Freeze header rows
        ExcelStyleHelper.freeze_header_rows(ws, num_rows=3)
    

    def _export_individual_files(self, object_names: List[str], output_path: str, 
                                stats: Dict) -> Tuple[str, Dict]:
        """
        Export each object to its own Excel file, auto-zip if multiple objects
        
        Args:
            object_names: List of object API names
            output_path: Base output path (will be modified for zip if needed)
            stats: Statistics dictionary
            
        Returns:
            Tuple of (final_output_path, stats)
                - If 1 object: returns .xlsx path
                - If 2+ objects: returns .zip path
        """
        self._log_status("📦 Individual Files Mode: Separate .xlsx per object")
        
        # Sort objects alphabetically
        sorted_objects = sorted(object_names)
        
        # Determine output strategy
        export_type = "Metadata"  # Will be used in filenames
        
        # Get base directory and filename
        output_dir = os.path.dirname(output_path)
        base_filename = os.path.splitext(os.path.basename(output_path))[0]
        
        # List to store created file paths
        created_files = []
        
        # Process each object
        for i, obj_name in enumerate(sorted_objects, 1):
            self._log_status(f"[{i}/{len(sorted_objects)}] Processing object: {obj_name}")
            
            try:
                fields = self._get_object_metadata(obj_name)
                stats['successful_objects'] += 1
                stats['total_fields'] += len(fields)
                
                # ✅ Get object label
                object_label = ExcelStyleHelper.get_object_label(self.sf_client, obj_name)
                
                # ✅ Create individual Excel file
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{obj_name}_{export_type}_{timestamp}.xlsx"
                file_path = os.path.join(output_dir, filename)
                
                self._create_individual_excel_file(
                    file_path,
                    obj_name,
                    object_label,
                    fields
                )
                
                created_files.append(file_path)
                
                self._log_status(
                    f"  ✅ Created: {filename} | Fields: {len(fields)}"
                )
                
            except Exception as e:
                error_msg = str(e)
                self._log_status(f"  ❌ ERROR: {error_msg}")
                stats['failed_objects'] += 1
                stats['failed_object_details'].append({'name': obj_name, 'reason': error_msg})
            
            self._log_status("")
        
        # ✅ Decide: Single file or ZIP?
        if len(created_files) == 0:
            # No files created
            self._log_status("⚠️ No files created - no data to export")
            raise Exception("No metadata found for any selected objects")
            
        elif len(created_files) == 1:
            # ✅ Single file - return as-is (no zip)
            final_path = created_files[0]
            self._log_status(f"✅ Single file created: {final_path}")
            return final_path, stats
            
        else:
            # ✅ Multiple files - create ZIP
            self._log_status(f"📦 Creating ZIP archive for {len(created_files)} files...")
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            zip_filename = f"Salesforce_{export_type}_Export_{timestamp}.zip"
            zip_path = os.path.join(output_dir, zip_filename)
            
            final_path = self._create_zip_archive(created_files, zip_path)
            
            return final_path, stats
    

    def _create_individual_excel_file(self, file_path: str, object_api: str, 
                                    object_label: str, fields: List[MetadataField]):
        """
        Create a single Excel file for one object with styling
        
        ✅ UPDATED: Use new METADATA_HEADERS
        """
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = ExcelStyleHelper.sanitize_sheet_name(object_label)
        
        # ✅ Use class constant for headers
        headers = self.METADATA_HEADERS
        num_cols = len(headers)
        
        # Add Title Row
        ExcelStyleHelper.add_title_row(
            ws,
            title=f"Salesforce Metadata Export - {object_label}",
            num_cols=num_cols,
            row_num=1
        )
        
        # Add Info Row with object details
        ExcelStyleHelper.add_info_row(
            ws,
            object_label=object_label,
            object_api=object_api,
            record_count=len(fields),
            num_cols=num_cols,
            row_num=2
        )
        
        # Add Header Row
        ExcelStyleHelper.add_header_row(ws, headers, row_num=3)
        
        # Add Data Rows with alternating colors
        for row_idx, field in enumerate(fields, start=4):
            row_data = field.to_row()
            is_even_row = (row_idx % 2 == 0)
            data_style = ExcelStyleHelper.get_data_style(is_even_row)
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                ExcelStyleHelper.apply_style_to_cell(cell, data_style)
        
        # Auto-adjust column widths
        ExcelStyleHelper.auto_adjust_column_widths(ws, headers)
        
        # Freeze header rows
        ExcelStyleHelper.freeze_header_rows(ws, num_rows=3)
        
        # Save workbook
        wb.save(file_path)
        
    
    def _create_zip_archive(self, file_paths: List[str], zip_path: str) -> str:
        """
        Create ZIP archive containing multiple Excel files
        
        Args:
            file_paths: List of file paths to include in ZIP
            zip_path: Output ZIP file path
            
        Returns:
            Path to created ZIP file
        """
        try:
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for file_path in file_paths:
                    # Add file to ZIP with just the filename (no directory structure)
                    arcname = os.path.basename(file_path)
                    zipf.write(file_path, arcname=arcname)
                    self._log_status(f"  📄 Added to ZIP: {arcname}")
            
            # ✅ Delete individual files after zipping
            self._log_status("🧹 Cleaning up individual files...")
            for file_path in file_paths:
                try:
                    os.remove(file_path)
                except Exception as e:
                    self._log_status(f"  ⚠️ Could not delete {file_path}: {e}")
            
            self._log_status(f"✅ ZIP archive created: {zip_path}")
            self._log_status(f"✅ Total files in ZIP: {len(file_paths)}")
            
            return zip_path
            
        except Exception as e:
            self._log_status(f"❌ Error creating ZIP: {str(e)}")
            raise   
        
    
    def _get_object_metadata(self, object_name: str) -> List[MetadataField]:
        """
        Get metadata for all fields of an object
        
        ✅ FIXED: Better error handling and parameter passing
        """
        metadata_fields = []
        
        try:
            obj_describe = getattr(self.sf, object_name).describe()
            
            for field in obj_describe['fields']:
                try:
                    # Extract field information
                    field_label = field.get('label', '')
                    api_name = field.get('name', '')
                    
                    # ✅ Get raw data type
                    data_type = field.get('type', '').capitalize()
                    
                    # ✅ Get field length
                    length = self._get_field_length(field)
                    
                    # ✅ Get detailed field type
                    field_type = self._format_field_type(field)
                    
                    # ✅ Check if required
                    required = self._is_field_required(field)
                    
                    # ✅ Get picklist values (if applicable)
                    picklist_values = self._get_picklist_values_string(field)
                    
                    # Formula
                    formula = field.get('calculatedFormula', '')
                    
                    # ✅ Check if external ID
                    external_id = 'Yes' if field.get('externalId', False) else ''
                    
                    # ✅ Check if history tracking enabled
                    track_history = 'Yes' if field.get('trackHistory', False) else ''
                    
                    # ✅ Get field description (from inlineHelpText)
                    description = field.get('inlineHelpText', '')
                    
                    # Help text (same as description in most cases)
                    help_text = field.get('inlineHelpText', '')
                    
                    # Attributes
                    attributes = self._get_field_attributes(field)
                    
                    # Field usage - wrap in try/catch to prevent blocking
                    try:
                        field_usage = self.usage_tracker.get_field_usage(object_name, api_name)
                    except Exception as usage_error:
                        self._log_status(f"    ⚠️ Could not get field usage for {api_name}: {str(usage_error)}")
                        field_usage = ""
                    
                    # Create MetadataField object
                    metadata_field = MetadataField(
                        object_name=object_name,
                        field_label=field_label,
                        api_name=api_name,
                        data_type=data_type,
                        length=length,
                        field_type=field_type,
                        required=required,
                        picklist_values=picklist_values,
                        formula=formula,
                        external_id=external_id,
                        track_history=track_history,
                        description=description,
                        help_text=help_text,
                        attributes=attributes,
                        field_usage=field_usage
                    )
                    
                    metadata_fields.append(metadata_field)
                    
                except Exception as field_error:
                    # Log but continue processing other fields
                    self._log_status(f"    ⚠️ Error processing field {field.get('name', 'unknown')}: {str(field_error)}")
                    continue
        
        except Exception as e:
            self._log_status(f"  ❌ ERROR in _get_object_metadata for {object_name}: {str(e)}")
            raise
        
        return metadata_fields    
    
    
    

    
    def _get_field_length(self, field: dict) -> str:
        """
        ✅ NEW METHOD: Get field length/size
        
        Returns:
            Field length as string (e.g., "255", "18,0" for precision/scale)
        """
        field_type = field.get('type', '')
        
        # Text fields - return length
        if field_type in ['string', 'textarea', 'url', 'email', 'phone', 'encryptedstring']:
            length = field.get('length', 0)
            return str(length) if length else ''
        
        # Number/Currency fields - return precision,scale
        elif field_type in ['double', 'currency', 'percent']:
            precision = field.get('precision', 0)
            scale = field.get('scale', 0)
            if precision:
                return f"{precision},{scale}"
            return ''
        
        # Auto-number - return length
        elif field_type == 'autonumber':
            length = field.get('length', 0)
            return str(length) if length else ''
        
        return ''
    
    def _is_field_required(self, field: dict) -> str:
        """
        ✅ NEW METHOD: Check if field is required
        
        Returns:
            'Yes' if required, '' if optional
        """
        # A field is required if:
        # 1. nillable = False (cannot be null)
        # 2. AND defaultedOnCreate = False (no default value)
        # 3. AND not a formula field (calculated fields can't be required)
        
        is_nillable = field.get('nillable', True)
        has_default = field.get('defaultedOnCreate', False)
        is_calculated = field.get('calculated', False)
        
        if not is_nillable and not has_default and not is_calculated:
            return 'Yes'
        
        return ''
    
    def _get_picklist_values_string(self, field: dict) -> str:
        """
        ✅ NEW METHOD: Get picklist values as comma-separated string
        
        Returns:
            Comma-separated list of active picklist values
        """
        field_type = field.get('type', '')
        
        if field_type not in ['picklist', 'multipicklist']:
            return ''
        
        try:
            picklist_values = field.get('picklistValues', [])
            
            # Get only active values
            active_values = [
                pv.get('label', pv.get('value', ''))
                for pv in picklist_values
                if pv.get('active', True)
            ]
            
            # Limit to first 10 values to avoid overly long cells
            if len(active_values) > 10:
                display_values = active_values[:10]
                return ', '.join(display_values) + f' ...and {len(active_values) - 10} more'
            else:
                return ', '.join(active_values)
                
        except Exception as e:
            return ''

    
    def _format_field_type(self, field: dict) -> str:
        """
        ✅ ENHANCED: Format field type with additional details
        
        Returns more detailed type information than before
        """
        field_type = field.get('type', '')
        
        # Text fields
        if field_type in ['string', 'textarea', 'url', 'email', 'phone']:
            length = field.get('length', 0)
            if length:
                return f"{field_type.capitalize()} ({length})"
        
        # Encrypted text
        elif field_type == 'encryptedstring':
            length = field.get('length', 0)
            return f"Encrypted Text ({length})" if length else "Encrypted Text"
        
        # Number/Currency fields
        elif field_type in ['double', 'currency', 'percent']:
            precision = field.get('precision', 0)
            scale = field.get('scale', 0)
            if precision:
                return f"{field_type.capitalize()} ({precision},{scale})"
        
        # Lookup/Master-Detail
        elif field_type == 'reference':
            ref_to = field.get('referenceTo', [])
            relationship_name = field.get('relationshipName', '')
            
            # Check if master-detail
            if field.get('cascadeDelete', False):
                rel_type = "Master-Detail"
            else:
                rel_type = "Lookup"
            
            if ref_to:
                ref_str = ', '.join(ref_to)
                if relationship_name:
                    return f"{rel_type} ({ref_str}) [{relationship_name}]"
                else:
                    return f"{rel_type} ({ref_str})"
        
        # Picklist
        elif field_type in ['picklist', 'multipicklist']:
            picklist_values = field.get('picklistValues', [])
            value_count = len([pv for pv in picklist_values if pv.get('active', True)])
            
            # Check if global picklist
            value_set = field.get('valueSet', {})
            value_set_name = value_set.get('valueSetName')
            
            if value_set_name:
                return f"{field_type.capitalize()} - Global ({value_count} values)"
            else:
                return f"{field_type.capitalize()} ({value_count} values)"
        
        # Auto-Number
        elif field_type == 'autonumber':
            display_format = field.get('autoNumber', {}).get('displayFormat', '')
            if display_format:
                return f"Auto Number ({display_format})"
            return "Auto Number"
        
        # Checkbox
        elif field_type == 'boolean':
            return "Checkbox"
        
        # Date/Time
        elif field_type == 'date':
            return "Date"
        elif field_type == 'datetime':
            return "Date/Time"
        elif field_type == 'time':
            return "Time"
        
        # Long Text Area
        elif field_type == 'textarea':
            length = field.get('length', 0)
            return f"Long Text Area ({length})" if length else "Long Text Area"
        
        # Rich Text Area
        elif field_type == 'richtextarea':
            length = field.get('length', 0)
            return f"Rich Text Area ({length})" if length else "Rich Text Area"
        
        # Default: capitalize type
        return field_type.capitalize()
    

    
    def _get_field_attributes(self, field: dict) -> str:
        """
        Get field attributes like Required, Unique, External ID, etc.
        
        ✅ UPDATED: Remove redundant attributes (now in dedicated columns)
        """
        attributes = []
        
        # Note: Required, External ID, Track History now have dedicated columns
        # So we remove them from attributes
        
        if field.get('unique', False):
            attributes.append('Unique')
        
        if field.get('autoNumber', False):
            attributes.append('Auto Number')
        
        if field.get('calculated', False):
            attributes.append('Formula')
        
        if field.get('cascadeDelete', False):
            attributes.append('Cascade Delete')
        
        if field.get('restrictedPicklist', False):
            attributes.append('Restricted Picklist')
        
        if field.get('encrypted', False):
            attributes.append('Encrypted')
        
        if field.get('dependentPicklist', False):
            attributes.append('Dependent Picklist')
        
        if field.get('filteredLookupInfo'):
            attributes.append('Filtered Lookup')
        
        return ', '.join(attributes) if attributes else ''


    
    def _create_csv_file(self, metadata_fields: List[MetadataField], output_path: str) -> str:
        """
        ⚠️ LEGACY: Create CSV file with metadata (old method)
        """
        # ✅ Use class constant for headers
        headers = self.METADATA_HEADERS
        
        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(headers)
            
            for field in metadata_fields:
                writer.writerow(field.to_row())
        
        self._log_status(f"✅ CSV file created: {output_path}")
        self._log_status(f"✅ Total fields exported: {len(metadata_fields)}")
        return output_path
    
    def _log_status(self, message: str):
        """Log status message"""
        if self.sf_client.status_callback:
            self.sf_client.status_callback(message, verbose=True)
