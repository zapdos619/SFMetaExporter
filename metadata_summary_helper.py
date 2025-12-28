"""
Metadata Summary Tab Helper
Generates summary statistics for metadata exports
"""
from typing import Dict, List, Tuple
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from excel_style_helper import ExcelStyleHelper
from datetime import datetime
from models import MetadataField


class MetadataSummaryData:
    """Stores summary statistics for a single object"""
    
    def __init__(self, object_name: str, object_api: str):
        self.object_name = object_name
        self.object_api = object_api
        self.master_object = ""  # Will be populated if this is a detail object
        self.object_type = "Custom" if object_api.endswith("__c") else "Standard"
        self.standard_field_count = 0
        self.custom_field_count = 0
    
    def to_row(self, serial_number: int) -> List:
        """Convert to Excel row format"""
        return [
            serial_number,
            self.object_name,
            self.object_api,
            self.master_object,
            self.object_type,
            self.standard_field_count,
            self.custom_field_count
        ]


class MetadataSummaryHelper:
    """Helper class for creating metadata summary tabs"""
    
    # Summary tab headers
    SUMMARY_HEADERS = [
        'SL',
        'Object Name',
        'Object API',
        'Master Object',
        'Object Type',
        'Standard Field',
        'Custom Field'
    ]
    
    @staticmethod
    def create_summary_sheet(wb: Workbook, summary_data: List[MetadataSummaryData], 
                           total_stats: Dict) -> Worksheet:
        """
        Create summary sheet as first tab in workbook
        
        Args:
            wb: Workbook instance
            summary_data: List of MetadataSummaryData objects
            total_stats: Dictionary with overall statistics
            
        Returns:
            Created worksheet
        """
        # Create sheet at the beginning
        ws = wb.create_sheet("Summary", 0)
        
        headers = MetadataSummaryHelper.SUMMARY_HEADERS
        num_cols = len(headers)
        
        # ✅ Add Title Row (Row 1)
        ExcelStyleHelper.add_title_row(
            ws,
            title="Salesforce Metadata Export - Summary",
            num_cols=num_cols,
            row_num=1
        )
        
        # ✅ Add Info Row (Row 2)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
        info_cell = ws.cell(row=2, column=1)
        export_date = datetime.now().strftime('%Y-%m-%d %H:%M')
        
        total_objects = len(summary_data)
        total_fields = total_stats.get('total_fields', 0)
        
        info_cell.value = (
            f"Total Objects: {total_objects} | "
            f"Total Fields: {total_fields} | "
            f"Export Date: {export_date}"
        )
        
        # Apply info style
        info_style = ExcelStyleHelper.get_info_style()
        ExcelStyleHelper.apply_style_to_cell(info_cell, info_style)
        for col_num in range(2, num_cols + 1):
            cell = ws.cell(row=2, column=col_num)
            ExcelStyleHelper.apply_style_to_cell(cell, info_style)
        ws.row_dimensions[2].height = 20
        
        # ✅ Add Header Row (Row 3)
        ExcelStyleHelper.add_header_row(ws, headers, row_num=3)
        
        # ✅ Add Data Rows with alternating colors
        for idx, summary_obj in enumerate(summary_data, start=1):
            row_num = idx + 3  # Start from row 4
            row_data = summary_obj.to_row(idx)
            
            is_even_row = (row_num % 2 == 0)
            data_style = ExcelStyleHelper.get_data_style(is_even_row)
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.value = value
                ExcelStyleHelper.apply_style_to_cell(cell, data_style)
        
        # ✅ Add Totals Row (if data exists)
        if summary_data:
            totals_row_num = len(summary_data) + 4
            
            # Calculate totals
            total_standard = sum(s.standard_field_count for s in summary_data)
            total_custom = sum(s.custom_field_count for s in summary_data)
            
            totals_data = [
                "",  # SL
                "TOTAL",  # Object Name
                "",  # Object API
                "",  # Master Object
                "",  # Object Type
                total_standard,
                total_custom
            ]
            
            # Apply bold style to totals row
            from openpyxl.styles import Font, PatternFill
            totals_style = {
                'font': Font(name='Calibri', size=11, bold=True),
                'fill': PatternFill(
                    start_color="E0E0E0",
                    end_color="E0E0E0",
                    fill_type="solid"
                ),
                'alignment': ExcelStyleHelper.get_data_style()['alignment'],
                'border': ExcelStyleHelper.get_data_style()['border']
            }
            
            for col_idx, value in enumerate(totals_data, start=1):
                cell = ws.cell(row=totals_row_num, column=col_idx)
                cell.value = value
                ExcelStyleHelper.apply_style_to_cell(cell, totals_style)
        
        # ✅ Auto-adjust column widths
        ExcelStyleHelper.auto_adjust_column_widths(ws, headers, max_width=40)
        
        # ✅ Freeze header rows
        ExcelStyleHelper.freeze_header_rows(ws, num_rows=3)
        
        return ws
    
    @staticmethod
    def analyze_metadata(object_api: str, fields: List[MetadataField], 
                        sf_client) -> MetadataSummaryData:
        """
        Analyze metadata for an object and generate summary statistics
        
        Args:
            object_api: Object API name
            fields: List of MetadataField objects for this object
            sf_client: SalesforceClient instance for getting object label
            
        Returns:
            MetadataSummaryData object
        """
        # Get object label
        object_label = ExcelStyleHelper.get_object_label(sf_client, object_api)
        
        # Create summary data object
        summary = MetadataSummaryData(object_label, object_api)
        
        # Count standard vs custom fields
        for field in fields:
            if field.api_name.endswith('__c'):
                summary.custom_field_count += 1
            else:
                summary.standard_field_count += 1
        
        # ✅ Get master object (if this is a detail in master-detail relationship)
        summary.master_object = MetadataSummaryHelper._get_master_object(
            sf_client, 
            object_api
        )
        
        return summary
    
    @staticmethod
    def _get_master_object(sf_client, object_api: str) -> str:
        """
        Get master object name if this object is a detail in a master-detail relationship
        
        Args:
            sf_client: SalesforceClient instance
            object_api: Object API name
            
        Returns:
            Comma-separated list of master object names, or empty string
        """
        try:
            obj_describe = getattr(sf_client.sf, object_api).describe()
            
            master_objects = []
            
            # Look for master-detail relationships
            for field in obj_describe['fields']:
                # Master-detail fields have:
                # 1. type = 'reference'
                # 2. cascadeDelete = True
                if field.get('type') == 'reference' and field.get('cascadeDelete', False):
                    # Get the referenced object(s)
                    ref_to = field.get('referenceTo', [])
                    for ref_obj in ref_to:
                        if ref_obj not in master_objects:
                            # Get label for the master object
                            try:
                                master_label = ExcelStyleHelper.get_object_label(
                                    sf_client, 
                                    ref_obj
                                )
                                master_objects.append(master_label)
                            except:
                                master_objects.append(ref_obj)
            
            # Return comma-separated list
            return ', '.join(master_objects) if master_objects else ''
            
        except Exception as e:
            # If error, return empty string
            return ''