
"""
Picklist export functionality
UPDATED: Added IsGlobal? column detection and updated headers
"""
import requests
from typing import List, Dict, Optional, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import zipfile
import os
from datetime import datetime
import tempfile

from config import API_VERSION
from models import FieldInfo, PicklistValueDetail, ProcessingResult
from salesforce_client import SalesforceClient
from excel_style_helper import ExcelStyleHelper
from picklist_summary_helper import PicklistSummaryHelper, PicklistSummaryData


class PicklistExporter:
    """Handles picklist data export from Salesforce"""
    
    def __init__(self, sf_client: SalesforceClient):
        """Initialize with Salesforce client"""
        self.sf_client = sf_client
        self.sf = sf_client.sf
        self.base_url = sf_client.base_url
        self.headers = sf_client.headers
        self.api_version = sf_client.api_version
    
    # âœ… UPDATED: New headers with "IsGlobal?" column
    PICKLIST_HEADERS = [
        'Object',
        'Field Label',
        'Field API',
        'Picklist Value Label',
        'Picklist Value API',
        'Status',
        'IsGlobal?'  # âœ… NEW COLUMN
    ]
    
    
    def export_picklists(self, object_names: List[str], output_path: str) -> Tuple[str, Dict]:
        """
        Export picklist values for specified objects
        
        âš ï¸ LEGACY METHOD: This uses the old Excel format without styling.
        New code should use export_picklists_excel() instead.
        
        Args:
            object_names: List of object API names
            output_path: Output Excel file path
            
        Returns:
            Tuple of (output_path, statistics_dict)
        """
        self._log_status("=== Starting Picklist Export ===")
        self._log_status(f"Total objects to process: {len(object_names)}")
        
        stats = {
            'total_objects': len(object_names), 
            'successful_objects': 0, 
            'failed_objects': 0, 
            'objects_not_found': 0, 
            'objects_with_zero_picklists': 0, 
            'objects_with_picklists': 0, 
            'total_picklist_fields': 0, 
            'total_values': 0, 
            'total_active_values': 0, 
            'total_inactive_values': 0,
            'total_global_picklists': 0,  # âœ… NEW
            'failed_object_details': [], 
            'objects_without_picklists': [], 
            'objects_not_found_list': []
        }
        
        all_rows = []
        
        for i, obj_name in enumerate(object_names, 1):
            self._log_status(f"[{i}/{len(object_names)}] Processing object: {obj_name}")
            try:
                result = self._process_object(obj_name)
                
                if not result.object_exists:
                    stats['objects_not_found'] += 1
                    stats['objects_not_found_list'].append(obj_name)
                    stats['failed_object_details'].append({
                        'name': obj_name, 
                        'reason': 'Object does not exist in org'
                    })
                    self._log_status(f"  âš ï¸  Object not found in org")
                elif result.picklist_fields_count == 0:
                    stats['objects_with_zero_picklists'] += 1
                    stats['objects_without_picklists'].append(obj_name)
                    stats['successful_objects'] += 1
                    self._log_status(f"  â„¹ï¸  No picklist fields found")
                else:
                    stats['objects_with_picklists'] += 1
                    stats['successful_objects'] += 1
                    stats['total_picklist_fields'] += result.picklist_fields_count
                    stats['total_global_picklists'] += result.global_picklist_count  # âœ… NEW
                    all_rows.extend(result.rows)
                    stats['total_values'] += result.values_processed
                    stats['total_inactive_values'] += result.inactive_values
                    stats['total_active_values'] += (result.values_processed - result.inactive_values)
                    self._log_status(
                        f"  âœ… Fields: {result.picklist_fields_count}, "
                        f"Active: {result.values_processed - result.inactive_values}, "
                        f"Inactive: {result.inactive_values}, "
                        f"Global: {result.global_picklist_count}"  # âœ… NEW
                    )
            except Exception as e:
                error_msg = str(e)
                self._log_status(f"  âŒ ERROR: {error_msg}")
                stats['failed_objects'] += 1
                stats['failed_object_details'].append({
                    'name': obj_name, 
                    'reason': error_msg
                })
            self._log_status("")
        
        self._log_status("=== Creating Excel File ===")
        final_output_path = self._create_excel_file(all_rows, output_path)
        return final_output_path, stats


    
    
    def export_picklists_excel(self, object_names: List[str], output_path: str, 
                            export_mode: str = "single_tab") -> Tuple[str, Dict]:
        """
        Export picklist values for specified objects to Excel with styling
        
        âœ… UPDATED: Now includes Summary Tab
        
        Args:
            object_names: List of object API names
            output_path: Path for output file(s)
            export_mode: "single_tab", "multi_tab", or "individual_files"
            
        Returns:
            Tuple of (output_path, statistics_dict)
        """
        self._log_status("=== Starting Picklist Export (Excel Format) ===")
        self._log_status(f"Export Mode: {export_mode}")
        self._log_status(f"Total objects to process: {len(object_names)}")
        
        stats = {
            'total_objects': len(object_names),
            'successful_objects': 0,
            'failed_objects': 0,
            'objects_not_found': 0,
            'objects_with_zero_picklists': 0,
            'objects_with_picklists': 0,
            'total_picklist_fields': 0,
            'total_values': 0,
            'total_active_values': 0,
            'total_inactive_values': 0,
            'total_global_picklists': 0,
            'failed_object_details': [],
            'objects_without_picklists': [],
            'objects_not_found_list': [],
            'export_mode': export_mode
        }
        
        # Dispatch to appropriate export method based on mode
        if export_mode == "single_tab":
            return self._export_single_tab_with_summary(object_names, output_path, stats)
        elif export_mode == "multi_tab":
            return self._export_multi_tab_with_summary(object_names, output_path, stats)
        elif export_mode == "individual_files":
            return self._export_individual_files_with_summary(object_names, output_path, stats)
        else:
            raise ValueError(f"Invalid export mode: {export_mode}")
    

    def _export_single_tab_with_summary(self, object_names: List[str], 
                                       output_path: str, stats: Dict) -> Tuple[str, Dict]:
        """
        âœ… UPDATED: Single tab export with Summary tab
        """
        self._log_status("ðŸ“„ Single Tab Mode: All objects in one sheet")
        
        # Collect all data AND summary data
        all_rows = []
        summary_data_list = []
        
        for i, obj_name in enumerate(sorted(object_names), 1):
            self._log_status(f"[{i}/{len(object_names)}] Processing object: {obj_name}")
            
            try:
                result = self._process_object(obj_name)
                
                if not result.object_exists:
                    stats['objects_not_found'] += 1
                    stats['objects_not_found_list'].append(obj_name)
                    stats['failed_object_details'].append({
                        'name': obj_name,
                        'reason': 'Object does not exist in org'
                    })
                    self._log_status(f"  âš ï¸ Object not found in org")
                    continue
                    
                elif result.picklist_fields_count == 0:
                    stats['objects_with_zero_picklists'] += 1
                    stats['objects_without_picklists'].append(obj_name)
                    stats['successful_objects'] += 1
                    self._log_status(f"  â„¹ï¸ No picklist fields found")
                    continue
                    
                else:
                    stats['objects_with_picklists'] += 1
                    stats['successful_objects'] += 1
                    stats['total_picklist_fields'] += result.picklist_fields_count
                    stats['total_global_picklists'] += result.global_picklist_count
                    stats['total_values'] += result.values_processed
                    stats['total_inactive_values'] += result.inactive_values
                    stats['total_active_values'] += (result.values_processed - result.inactive_values)
                    
                    # Add rows to collection
                    all_rows.extend(result.rows)
                    
                    # âœ… Generate summary data for this object
                    summary_obj = PicklistSummaryHelper.analyze_picklist_data(
                        obj_name,
                        result.rows,
                        self.sf_client
                    )
                    summary_data_list.append(summary_obj)
                    
                    self._log_status(
                        f"  âœ… Fields: {result.picklist_fields_count}, "
                        f"Active: {result.values_processed - result.inactive_values}, "
                        f"Inactive: {result.inactive_values}"
                    )
                    
            except Exception as e:
                error_msg = str(e)
                self._log_status(f"  âŒ ERROR: {error_msg}")
                stats['failed_objects'] += 1
                stats['failed_object_details'].append({
                    'name': obj_name,
                    'reason': error_msg
                })
            
            self._log_status("")
        
        # Create Excel file with Summary tab + Data tab
        self._log_status("=== Creating Excel File with Summary ===")
        final_output_path = self._create_single_tab_excel_with_summary(
            all_rows, 
            summary_data_list, 
            output_path, 
            stats
        )
        
        return final_output_path, stats


    def _create_single_tab_excel_with_summary(self, rows: List[List[str]], 
                                             summary_data: List[PicklistSummaryData],
                                             output_path: str, stats: Dict) -> str:
        """
        âœ… NEW: Create Excel with Summary tab (first) + Data tab (second)
        """
        wb = Workbook()
        
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # âœ… Create Summary Tab (first)
        if summary_data:
            self._log_status("ðŸ“Š Creating Summary tab...")
            PicklistSummaryHelper.create_summary_sheet(wb, summary_data, stats)
        
        # âœ… Create Data Tab (second)
        self._log_status("ðŸ“‹ Creating Picklist Data tab...")
        ws_data = wb.create_sheet("Picklist Data")
        
        headers = self.PICKLIST_HEADERS
        num_cols = len(headers)
        
        # Add Title Row
        ExcelStyleHelper.add_title_row(
            ws_data,
            title="Salesforce Picklist Export",
            num_cols=num_cols,
            row_num=1
        )
        
        # Add Info Row
        total_objects = stats['successful_objects']
        total_records = len(rows)
        
        ws_data.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
        info_cell = ws_data.cell(row=2, column=1)
        export_date = datetime.now().strftime('%Y-%m-%d %H:%M')
        info_cell.value = (
            f"Objects: {total_objects} | "
            f"Total Picklist Values: {total_records} | "
            f"Global Picklists: {stats.get('total_global_picklists', 0)} | "
            f"Export Date: {export_date}"
        )
        
        info_style = ExcelStyleHelper.get_info_style()
        ExcelStyleHelper.apply_style_to_cell(info_cell, info_style)
        for col_num in range(2, num_cols + 1):
            cell = ws_data.cell(row=2, column=col_num)
            ExcelStyleHelper.apply_style_to_cell(cell, info_style)
        ws_data.row_dimensions[2].height = 20
        
        # Add Header Row
        ExcelStyleHelper.add_header_row(ws_data, headers, row_num=3)
        
        # Add Data Rows with alternating colors
        for row_idx, row_data in enumerate(rows, start=4):
            is_even_row = (row_idx % 2 == 0)
            data_style = ExcelStyleHelper.get_data_style(is_even_row)
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_data.cell(row=row_idx, column=col_idx)
                cell.value = value
                ExcelStyleHelper.apply_style_to_cell(cell, data_style)
        
        # Auto-adjust column widths
        ExcelStyleHelper.auto_adjust_column_widths(ws_data, headers)
        
        # Freeze header rows
        ExcelStyleHelper.freeze_header_rows(ws_data, num_rows=3)
        
        # Save workbook
        wb.save(output_path)
        
        self._log_status(f"âœ… Excel file created: {output_path}")
        self._log_status(f"âœ… Total sheets: {len(wb.worksheets)} (Summary + Data)")
        
        return output_path

        
    def _export_multi_tab_with_summary(self, object_names: List[str], 
                                      output_path: str, stats: Dict) -> Tuple[str, Dict]:
        """
        âœ… UPDATED: Multiple tabs export with Summary tab
        """
        self._log_status("ðŸ“‘ Multiple Tabs Mode: One sheet per object + Summary")
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # Sort objects alphabetically
        sorted_objects = sorted(object_names)
        
        # Store summary data
        summary_data_list = []
        
        # Process each object and create a sheet
        for i, obj_name in enumerate(sorted_objects, 1):
            self._log_status(f"[{i}/{len(sorted_objects)}] Processing object: {obj_name}")
            
            try:
                result = self._process_object(obj_name)
                
                if not result.object_exists:
                    stats['objects_not_found'] += 1
                    stats['objects_not_found_list'].append(obj_name)
                    stats['failed_object_details'].append({
                        'name': obj_name,
                        'reason': 'Object does not exist in org'
                    })
                    self._log_status(f"  âš ï¸ Object not found in org")
                    continue
                    
                elif result.picklist_fields_count == 0:
                    stats['objects_with_zero_picklists'] += 1
                    stats['objects_without_picklists'].append(obj_name)
                    stats['successful_objects'] += 1
                    self._log_status(f"  â„¹ï¸ No picklist fields found")
                    continue
                    
                else:
                    stats['objects_with_picklists'] += 1
                    stats['successful_objects'] += 1
                    stats['total_picklist_fields'] += result.picklist_fields_count
                    stats['total_global_picklists'] += result.global_picklist_count
                    stats['total_values'] += result.values_processed
                    stats['total_inactive_values'] += result.inactive_values
                    stats['total_active_values'] += (result.values_processed - result.inactive_values)
                    
                    # Get object label for tab name
                    object_label = ExcelStyleHelper.get_object_label(self.sf_client, obj_name)
                    
                    # Create sheet for this object
                    self._create_sheet_for_object(
                        wb, 
                        obj_name, 
                        object_label, 
                        result.rows, 
                        result.picklist_fields_count,
                        result.values_processed,
                        result.inactive_values
                    )
                    
                    # âœ… Generate summary data for this object
                    summary_obj = PicklistSummaryHelper.analyze_picklist_data(
                        obj_name,
                        result.rows,
                        self.sf_client
                    )
                    summary_data_list.append(summary_obj)
                    
                    self._log_status(
                        f"  âœ… Fields: {result.picklist_fields_count}, "
                        f"Active: {result.values_processed - result.inactive_values}, "
                        f"Inactive: {result.inactive_values}"
                    )
                    
            except Exception as e:
                error_msg = str(e)
                self._log_status(f"  âŒ ERROR: {error_msg}")
                stats['failed_objects'] += 1
                stats['failed_object_details'].append({
                    'name': obj_name,
                    'reason': error_msg
                })
            
            self._log_status("")
        
        # âœ… Create Summary Tab (first tab)
        if summary_data_list:
            self._log_status("ðŸ“Š Creating Summary tab...")
            PicklistSummaryHelper.create_summary_sheet(wb, summary_data_list, stats)
        else:
            # No data - create placeholder
            ws = wb.create_sheet("No Data", 0)
            ws['A1'] = "No picklist data found for selected objects"
            self._log_status("âš ï¸ No data exported - created placeholder sheet")
        
        # Save workbook
        wb.save(output_path)
        
        self._log_status(f"âœ… Excel file created: {output_path}")
        self._log_status(f"âœ… Total sheets: {len(wb.worksheets)} (Summary + {len(summary_data_list)} objects)")
        
        return output_path, stats

    def _export_individual_files_with_summary(self, object_names: List[str], 
                                             output_path: str, stats: Dict) -> Tuple[str, Dict]:
        """
        âœ… UPDATED: Individual files export with separate Summary file
        """
        self._log_status("ðŸ“¦ Individual Files Mode: Separate .xlsx per object + Summary file")
        
        # Sort objects alphabetically
        sorted_objects = sorted(object_names)
        
        # Determine output strategy
        export_type = "Picklist"
        
        # Get base directory and filename
        output_dir = os.path.dirname(output_path)
        base_filename = os.path.splitext(os.path.basename(output_path))[0]
        
        # List to store created file paths
        created_files = []
        summary_data_list = []
        
        # Process each object
        for i, obj_name in enumerate(sorted_objects, 1):
            self._log_status(f"[{i}/{len(sorted_objects)}] Processing object: {obj_name}")
            
            try:
                result = self._process_object(obj_name)
                
                if not result.object_exists:
                    stats['objects_not_found'] += 1
                    stats['objects_not_found_list'].append(obj_name)
                    stats['failed_object_details'].append({
                        'name': obj_name,
                        'reason': 'Object does not exist in org'
                    })
                    self._log_status(f"  âš ï¸ Object not found in org")
                    continue
                    
                elif result.picklist_fields_count == 0:
                    stats['objects_with_zero_picklists'] += 1
                    stats['objects_without_picklists'].append(obj_name)
                    stats['successful_objects'] += 1
                    self._log_status(f"  â„¹ï¸ No picklist fields found")
                    continue
                    
                else:
                    stats['objects_with_picklists'] += 1
                    stats['successful_objects'] += 1
                    stats['total_picklist_fields'] += result.picklist_fields_count
                    stats['total_global_picklists'] += result.global_picklist_count
                    stats['total_values'] += result.values_processed
                    stats['total_inactive_values'] += result.inactive_values
                    stats['total_active_values'] += (result.values_processed - result.inactive_values)
                    
                    # Get object label
                    object_label = ExcelStyleHelper.get_object_label(self.sf_client, obj_name)
                    
                    # Create individual Excel file
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = f"{obj_name}_{export_type}_{timestamp}.xlsx"
                    file_path = os.path.join(output_dir, filename)
                    
                    self._create_individual_excel_file(
                        file_path,
                        obj_name,
                        object_label,
                        result.rows,
                        result.picklist_fields_count,
                        result.values_processed,
                        result.inactive_values
                    )
                    
                    created_files.append(file_path)
                    
                    # âœ… Generate summary data
                    summary_obj = PicklistSummaryHelper.analyze_picklist_data(
                        obj_name,
                        result.rows,
                        self.sf_client
                    )
                    summary_data_list.append(summary_obj)
                    
                    self._log_status(
                        f"  âœ… Created: {filename} | "
                        f"Fields: {result.picklist_fields_count}, "
                        f"Values: {result.values_processed}"
                    )
                    
            except Exception as e:
                error_msg = str(e)
                self._log_status(f"  âŒ ERROR: {error_msg}")
                stats['failed_objects'] += 1
                stats['failed_object_details'].append({
                    'name': obj_name,
                    'reason': error_msg
                })
            
            self._log_status("")
        
        # âœ… Create Summary File
        if summary_data_list:
            self._log_status("ðŸ“Š Creating Summary file...")
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            summary_filename = f"Summary_{export_type}_{timestamp}.xlsx"
            summary_path = os.path.join(output_dir, summary_filename)
            
            wb_summary = Workbook()
            default_sheet = wb_summary.active
            wb_summary.remove(default_sheet)
            
            PicklistSummaryHelper.create_summary_sheet(wb_summary, summary_data_list, stats)
            wb_summary.save(summary_path)
            
            created_files.insert(0, summary_path)  # Add summary at beginning
            self._log_status(f"âœ… Summary file created: {summary_filename}")
        
        # Decide: Single file or ZIP?
        if len(created_files) == 0:
            self._log_status("âš ï¸ No files created - no data to export")
            raise Exception("No picklist data found for any selected objects")
            
        elif len(created_files) == 1:
            # Single file (just summary)
            final_path = created_files[0]
            self._log_status(f"âœ… Single file created: {final_path}")
            return final_path, stats
            
        else:
            # Multiple files - create ZIP
            self._log_status(f"ðŸ“¦ Creating ZIP archive for {len(created_files)} files...")
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            zip_filename = f"Salesforce_{export_type}_Export_{timestamp}.zip"
            zip_path = os.path.join(output_dir, zip_filename)
            
            final_path = self._create_zip_archive(created_files, zip_path)
            
            return final_path, stats

    
    
    def _export_single_tab(self, object_names: List[str], output_path: str, 
                        stats: Dict) -> Tuple[str, Dict]:
        """
        Export all objects to a single Excel sheet with styling
        
        Args:
            object_names: List of object API names
            output_path: Output Excel file path
            stats: Statistics dictionary
            
        Returns:
            Tuple of (output_path, stats)
        """
        self._log_status("ðŸ“„ Single Tab Mode: All objects in one sheet")
        
        # Collect all data first
        all_rows = []
        
        for i, obj_name in enumerate(sorted(object_names), 1):  # âœ… Sort alphabetically
            self._log_status(f"[{i}/{len(object_names)}] Processing object: {obj_name}")
            
            try:
                result = self._process_object(obj_name)
                
                if not result.object_exists:
                    stats['objects_not_found'] += 1
                    stats['objects_not_found_list'].append(obj_name)
                    stats['failed_object_details'].append({
                        'name': obj_name,
                        'reason': 'Object does not exist in org'
                    })
                    self._log_status(f"  âš ï¸ Object not found in org")
                    
                elif result.picklist_fields_count == 0:
                    stats['objects_with_zero_picklists'] += 1
                    stats['objects_without_picklists'].append(obj_name)
                    stats['successful_objects'] += 1
                    self._log_status(f"  â„¹ï¸ No picklist fields found")
                    
                else:
                    stats['objects_with_picklists'] += 1
                    stats['successful_objects'] += 1
                    stats['total_picklist_fields'] += result.picklist_fields_count
                    
                    # Add rows to collection
                    all_rows.extend(result.rows)
                    
                    stats['total_values'] += result.values_processed
                    stats['total_inactive_values'] += result.inactive_values
                    stats['total_active_values'] += (result.values_processed - result.inactive_values)
                    
                    self._log_status(
                        f"  âœ… Fields: {result.picklist_fields_count}, "
                        f"Active: {result.values_processed - result.inactive_values}, "
                        f"Inactive: {result.inactive_values}"
                    )
                    
            except Exception as e:
                error_msg = str(e)
                self._log_status(f"  âŒ ERROR: {error_msg}")
                stats['failed_objects'] += 1
                stats['failed_object_details'].append({
                    'name': obj_name,
                    'reason': error_msg
                })
            
            self._log_status("")
        
        # Create Excel file with styling
        self._log_status("=== Creating Excel File ===")
        final_output_path = self._create_single_tab_excel(all_rows, output_path, stats)
        
        return final_output_path, stats    
        
    
    def _create_single_tab_excel(self, rows: List[List[str]], output_path: str, 
                                stats: Dict) -> str:
        """
        Create Excel file with single tab and styling
        
        âœ… UPDATED: Use new PICKLIST_HEADERS
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Picklist Export"
        
        # âœ… Use class constant for headers
        headers = self.PICKLIST_HEADERS
        num_cols = len(headers)
        
        # âœ… Add Title Row (Row 1)
        ExcelStyleHelper.add_title_row(
            ws,
            title="Salesforce Picklist Export",
            num_cols=num_cols,
            row_num=1
        )
        
        # âœ… Add Info Row (Row 2)
        total_objects = stats['successful_objects']
        total_records = len(rows)
        
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
        info_cell = ws.cell(row=2, column=1)
        export_date = datetime.now().strftime('%Y-%m-%d %H:%M')
        info_cell.value = (
            f"Objects: {total_objects} | "
            f"Total Picklist Values: {total_records} | "
            f"Global Picklists: {stats.get('total_global_picklists', 0)} | "  # âœ… NEW
            f"Export Date: {export_date}"
        )
        
        # Apply info style
        info_style = ExcelStyleHelper.get_info_style()
        ExcelStyleHelper.apply_style_to_cell(info_cell, info_style)
        for col_num in range(2, num_cols + 1):
            cell = ws.cell(row=2, column=col_num)
            ExcelStyleHelper.apply_style_to_cell(cell, info_style)
        ws.row_dimensions[2].height = 20
        
        # âœ… Add Header Row (Row 3)
        ExcelStyleHelper.add_header_row(ws, headers, row_num=3)
        
        # âœ… Add Data Rows with alternating colors
        for row_idx, row_data in enumerate(rows, start=4):
            is_even_row = (row_idx % 2 == 0)
            data_style = ExcelStyleHelper.get_data_style(is_even_row)
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                ExcelStyleHelper.apply_style_to_cell(cell, data_style)
        
        # âœ… Auto-adjust column widths
        ExcelStyleHelper.auto_adjust_column_widths(ws, headers)
        
        # âœ… Freeze header rows
        ExcelStyleHelper.freeze_header_rows(ws, num_rows=3)
        
        # Save workbook
        wb.save(output_path)
        
        self._log_status(f"âœ… Excel file created: {output_path}")
        self._log_status(f"âœ… Total data rows: {len(rows)}")
        
        return output_path


        
    
    def _export_multi_tab(self, object_names: List[str], output_path: str, 
                        stats: Dict) -> Tuple[str, Dict]:
        """
        Export each object to its own sheet in a single Excel file
        
        Args:
            object_names: List of object API names
            output_path: Output Excel file path
            stats: Statistics dictionary
            
        Returns:
            Tuple of (output_path, stats)
        """
        self._log_status("ðŸ“‘ Multiple Tabs Mode: One sheet per object")
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Sort objects alphabetically
        sorted_objects = sorted(object_names)
        
        # Process each object and create a sheet
        for i, obj_name in enumerate(sorted_objects, 1):
            self._log_status(f"[{i}/{len(sorted_objects)}] Processing object: {obj_name}")
            
            try:
                result = self._process_object(obj_name)
                
                if not result.object_exists:
                    stats['objects_not_found'] += 1
                    stats['objects_not_found_list'].append(obj_name)
                    stats['failed_object_details'].append({
                        'name': obj_name,
                        'reason': 'Object does not exist in org'
                    })
                    self._log_status(f"  âš ï¸ Object not found in org")
                    continue
                    
                elif result.picklist_fields_count == 0:
                    stats['objects_with_zero_picklists'] += 1
                    stats['objects_without_picklists'].append(obj_name)
                    stats['successful_objects'] += 1
                    self._log_status(f"  â„¹ï¸ No picklist fields found")
                    continue
                    
                else:
                    stats['objects_with_picklists'] += 1
                    stats['successful_objects'] += 1
                    stats['total_picklist_fields'] += result.picklist_fields_count
                    stats['total_values'] += result.values_processed
                    stats['total_inactive_values'] += result.inactive_values
                    stats['total_active_values'] += (result.values_processed - result.inactive_values)
                    
                    # âœ… Get object label for tab name
                    object_label = ExcelStyleHelper.get_object_label(self.sf_client, obj_name)
                    
                    # âœ… Create sheet for this object
                    self._create_sheet_for_object(
                        wb, 
                        obj_name, 
                        object_label, 
                        result.rows, 
                        result.picklist_fields_count,
                        result.values_processed,
                        result.inactive_values
                    )
                    
                    self._log_status(
                        f"  âœ… Fields: {result.picklist_fields_count}, "
                        f"Active: {result.values_processed - result.inactive_values}, "
                        f"Inactive: {result.inactive_values}"
                    )
                    
            except Exception as e:
                error_msg = str(e)
                self._log_status(f"  âŒ ERROR: {error_msg}")
                stats['failed_objects'] += 1
                stats['failed_object_details'].append({
                    'name': obj_name,
                    'reason': error_msg
                })
            
            self._log_status("")
        
        # Save workbook
        if len(wb.worksheets) == 0:
            # No sheets created - create a placeholder
            ws = wb.create_sheet("No Data")
            ws['A1'] = "No picklist data found for selected objects"
            self._log_status("âš ï¸ No data exported - created placeholder sheet")
        
        wb.save(output_path)
        
        self._log_status(f"âœ… Excel file created: {output_path}")
        self._log_status(f"âœ… Total sheets: {len(wb.worksheets)}")
        
        return output_path, stats    
    
    
    def _create_sheet_for_object(
        self, 
        wb: Workbook, 
        object_api: str, 
        object_label: str, 
        rows: List[List[str]], 
        field_count: int, 
        total_values: int, 
        inactive_values: int
    ):
        """
        Create a single sheet for an object with styling
        
        âœ… UPDATED: Use new PICKLIST_HEADERS
        """
        # Sanitize sheet name
        sheet_name = ExcelStyleHelper.sanitize_sheet_name(object_label)
        
        # Create sheet
        ws = wb.create_sheet(sheet_name)
        
        # âœ… Use class constant for headers
        headers = self.PICKLIST_HEADERS
        num_cols = len(headers)
        
        # âœ… Add Title Row
        ExcelStyleHelper.add_title_row(
            ws,
            title=f"Salesforce Picklist Export - {object_label}",
            num_cols=num_cols,
            row_num=1
        )
        
        # âœ… Add Info Row
        ExcelStyleHelper.add_info_row(
            ws,
            object_label=object_label,
            object_api=object_api,
            record_count=total_values,
            num_cols=num_cols,
            row_num=2
        )
        
        # âœ… Add Header Row
        ExcelStyleHelper.add_header_row(ws, headers, row_num=3)
        
        # âœ… Add Data Rows with alternating colors
        for row_idx, row_data in enumerate(rows, start=4):
            is_even_row = (row_idx % 2 == 0)
            data_style = ExcelStyleHelper.get_data_style(is_even_row)
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                ExcelStyleHelper.apply_style_to_cell(cell, data_style)
        
        # âœ… Auto-adjust column widths
        ExcelStyleHelper.auto_adjust_column_widths(ws, headers)
        
        # âœ… Freeze header rows
        ExcelStyleHelper.freeze_header_rows(ws, num_rows=3)
    
    def _create_individual_excel_file(
        self, 
        file_path: str, 
        object_api: str, 
        object_label: str, 
        rows: List[List[str]], 
        field_count: int, 
        total_values: int, 
        inactive_values: int
    ):
        """
        Create a single Excel file for one object with styling
        
        âœ… UPDATED: Use new PICKLIST_HEADERS
        """
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = ExcelStyleHelper.sanitize_sheet_name(object_label)
        
        # âœ… Use class constant for headers
        headers = self.PICKLIST_HEADERS
        num_cols = len(headers)
        
        # âœ… Add Title Row
        ExcelStyleHelper.add_title_row(
            ws,
            title=f"Salesforce Picklist Export - {object_label}",
            num_cols=num_cols,
            row_num=1
        )
        
        # âœ… Add Info Row
        ExcelStyleHelper.add_info_row(
            ws,
            object_label=object_label,
            object_api=object_api,
            record_count=total_values,
            num_cols=num_cols,
            row_num=2
        )
        
        # âœ… Add Header Row
        ExcelStyleHelper.add_header_row(ws, headers, row_num=3)
        
        # âœ… Add Data Rows with alternating colors
        for row_idx, row_data in enumerate(rows, start=4):
            is_even_row = (row_idx % 2 == 0)
            data_style = ExcelStyleHelper.get_data_style(is_even_row)
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                ExcelStyleHelper.apply_style_to_cell(cell, data_style)
        
        # âœ… Auto-adjust column widths
        ExcelStyleHelper.auto_adjust_column_widths(ws, headers)
        
        # âœ… Freeze header rows
        ExcelStyleHelper.freeze_header_rows(ws, num_rows=3)
        
        # Save workbook
        wb.save(file_path)

    
    def _export_individual_files(self, object_names: List[str], output_path: str, 
                                stats: Dict) -> Tuple[str, Dict]:
        """
        Export each object to its own Excel file, auto-zip if multiple objects
        
        Args:
            object_names: List of object API names
            output_path: Base output path (will be modified for zip if needed)
            stats: Statistics dictionary
            
        Returns:
            Tuple of (final_output_path, stats)
                - If 1 object: returns .xlsx path
                - If 2+ objects: returns .zip path
        """
        self._log_status("ðŸ“¦ Individual Files Mode: Separate .xlsx per object")
        
        # Sort objects alphabetically
        sorted_objects = sorted(object_names)
        
        # Determine output strategy
        export_type = "Picklist"  # Will be used in filenames
        
        # Get base directory and filename
        output_dir = os.path.dirname(output_path)
        base_filename = os.path.splitext(os.path.basename(output_path))[0]
        
        # List to store created file paths
        created_files = []
        
        # Process each object
        for i, obj_name in enumerate(sorted_objects, 1):
            self._log_status(f"[{i}/{len(sorted_objects)}] Processing object: {obj_name}")
            
            try:
                result = self._process_object(obj_name)
                
                if not result.object_exists:
                    stats['objects_not_found'] += 1
                    stats['objects_not_found_list'].append(obj_name)
                    stats['failed_object_details'].append({
                        'name': obj_name,
                        'reason': 'Object does not exist in org'
                    })
                    self._log_status(f"  âš ï¸ Object not found in org")
                    continue
                    
                elif result.picklist_fields_count == 0:
                    stats['objects_with_zero_picklists'] += 1
                    stats['objects_without_picklists'].append(obj_name)
                    stats['successful_objects'] += 1
                    self._log_status(f"  â„¹ï¸ No picklist fields found")
                    continue
                    
                else:
                    stats['objects_with_picklists'] += 1
                    stats['successful_objects'] += 1
                    stats['total_picklist_fields'] += result.picklist_fields_count
                    stats['total_values'] += result.values_processed
                    stats['total_inactive_values'] += result.inactive_values
                    stats['total_active_values'] += (result.values_processed - result.inactive_values)
                    
                    # âœ… Get object label
                    object_label = ExcelStyleHelper.get_object_label(self.sf_client, obj_name)
                    
                    # âœ… Create individual Excel file
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = f"{obj_name}_{export_type}_{timestamp}.xlsx"
                    file_path = os.path.join(output_dir, filename)
                    
                    self._create_individual_excel_file(
                        file_path,
                        obj_name,
                        object_label,
                        result.rows,
                        result.picklist_fields_count,
                        result.values_processed,
                        result.inactive_values
                    )
                    
                    created_files.append(file_path)
                    
                    self._log_status(
                        f"  âœ… Created: {filename} | "
                        f"Fields: {result.picklist_fields_count}, "
                        f"Values: {result.values_processed}"
                    )
                    
            except Exception as e:
                error_msg = str(e)
                self._log_status(f"  âŒ ERROR: {error_msg}")
                stats['failed_objects'] += 1
                stats['failed_object_details'].append({
                    'name': obj_name,
                    'reason': error_msg
                })
            
            self._log_status("")
        
        # âœ… Decide: Single file or ZIP?
        if len(created_files) == 0:
            # No files created
            self._log_status("âš ï¸ No files created - no data to export")
            raise Exception("No picklist data found for any selected objects")
            
        elif len(created_files) == 1:
            # âœ… Single file - return as-is (no zip)
            final_path = created_files[0]
            self._log_status(f"âœ… Single file created: {final_path}")
            return final_path, stats
            
        else:
            # âœ… Multiple files - create ZIP
            self._log_status(f"ðŸ“¦ Creating ZIP archive for {len(created_files)} files...")
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            zip_filename = f"Salesforce_{export_type}_Export_{timestamp}.zip"
            zip_path = os.path.join(output_dir, zip_filename)
            
            final_path = self._create_zip_archive(created_files, zip_path)
            
            return final_path, stats
    
    def _create_individual_excel_file(self, file_path: str, object_api: str, 
                                    object_label: str, rows: List[List[str]], 
                                    field_count: int, total_values: int, 
                                    inactive_values: int):
        """
        Create a single Excel file for one object with styling
        
        Args:
            file_path: Full path for output Excel file
            object_api: Object API name
            object_label: Object label (for display)
            rows: Data rows for this object
            field_count: Number of picklist fields
            total_values: Total picklist values
            inactive_values: Number of inactive values
        """
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = ExcelStyleHelper.sanitize_sheet_name(object_label)
        
        # Define headers (same as current CSV columns)
        headers = [
            'Object',
            'Field Label',
            'Field API',
            'Picklist Value Label',
            'Picklist Value API',
            'Status'
        ]
        
        num_cols = len(headers)
        
        # âœ… Add Title Row (Row 1)
        ExcelStyleHelper.add_title_row(
            ws,
            title=f"Salesforce Picklist Export - {object_label}",
            num_cols=num_cols,
            row_num=1
        )
        
        # âœ… Add Info Row (Row 2) with object details
        ExcelStyleHelper.add_info_row(
            ws,
            object_label=object_label,
            object_api=object_api,
            record_count=total_values,
            num_cols=num_cols,
            row_num=2
        )
        
        # âœ… Add Header Row (Row 3)
        ExcelStyleHelper.add_header_row(ws, headers, row_num=3)
        
        # âœ… Add Data Rows (Starting from Row 4)
        data_style = ExcelStyleHelper.get_data_style()
        
        for row_idx, row_data in enumerate(rows, start=4):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                ExcelStyleHelper.apply_style_to_cell(cell, data_style)
        
        # âœ… Auto-adjust column widths
        ExcelStyleHelper.auto_adjust_column_widths(ws, headers)
        
        # âœ… Freeze header rows (title, info, headers)
        ExcelStyleHelper.freeze_header_rows(ws, num_rows=3)
        
        # Save workbook
        wb.save(file_path)
    
    
    def _create_zip_archive(self, file_paths: List[str], zip_path: str) -> str:
        """
        Create ZIP archive containing multiple Excel files
        
        Args:
            file_paths: List of file paths to include in ZIP
            zip_path: Output ZIP file path
            
        Returns:
            Path to created ZIP file
        """
        try:
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for file_path in file_paths:
                    arcname = os.path.basename(file_path)
                    zipf.write(file_path, arcname=arcname)
                    self._log_status(f"  ðŸ“„ Added to ZIP: {arcname}")
            
            # Delete individual files after zipping
            self._log_status("ðŸ§¹ Cleaning up individual files...")
            for file_path in file_paths:
                try:
                    os.remove(file_path)
                except Exception as e:
                    self._log_status(f"  âš ï¸ Could not delete {file_path}: {e}")
            
            self._log_status(f"âœ… ZIP archive created: {zip_path}")
            self._log_status(f"âœ… Total files in ZIP: {len(file_paths)}")
            
            return zip_path
            
        except Exception as e:
            self._log_status(f"âŒ Error creating ZIP: {str(e)}")
            raise

    
    
    def _process_object(self, obj_name: str) -> ProcessingResult:
        """
        Process a single object for picklist fields
        
        âœ… UPDATED: Tracks global picklists and adds IsGlobal column
        """
        result = ProcessingResult()
        
        try:
            getattr(self.sf, obj_name).describe()
        except Exception as e:
            if 'NOT_FOUND' in str(e) or 'INVALID_TYPE' in str(e):
                result.object_exists = False
                return result
            raise
        
        picklist_fields = self._get_picklist_fields(obj_name)
        result.picklist_fields_count = len(picklist_fields)
        
        if not picklist_fields:
            return result
        
        self._log_status(f"  Found {len(picklist_fields)} picklist fields")
        
        entity_def_id = self._resolve_entity_definition_id(obj_name)
        if entity_def_id:
            self._log_status(f"  EntityDefinition.Id: {entity_def_id}")
        
        for field_api, field_info in picklist_fields.items():
            # âœ… Track global picklists
            if field_info.is_global:
                result.global_picklist_count += 1
            
            values = self._query_picklist_values_with_fallback(
                obj_name, 
                entity_def_id, 
                field_api
            )
            
            if not values:
                continue
            
            self._log_status(
                f"    Field: {field_api} - {len(values)} values "
                f"{'(Global)' if field_info.is_global else ''}"  # âœ… NEW
            )
            
            for value in values:
                is_active = value.is_active if value.is_active is not None else True
                status = 'Active' if is_active else 'Inactive'
                
                if not is_active:
                    result.inactive_values += 1
                
                # âœ… UPDATED: Add IsGlobal column (7th column)
                row = [
                    obj_name,                  # Object
                    field_info.label,          # Field Label
                    field_api,                 # Field API
                    value.label,               # Picklist Value Label
                    value.value,               # Picklist Value API
                    status,                    # Status
                    'Yes' if field_info.is_global else ''  # âœ… IsGlobal? (NEW)
                ]
                
                result.rows.append(row)
                result.values_processed += 1
        
        return result    
    
    

    
    
    def _get_picklist_fields(self, object_name: str) -> Dict[str, FieldInfo]:
        """
        Get all picklist fields for an object
        
        âœ… UPDATED: Now detects global picklists
        """
        fields_dict = {}
        try:
            obj_describe = getattr(self.sf, object_name).describe()
            
            for field in obj_describe['fields']:
                if field['type'] in ['picklist', 'multipicklist']:
                    # âœ… NEW: Check if field uses global value set
                    is_global = self._is_global_picklist(field)
                    
                    fields_dict[field['name']] = FieldInfo(
                        api_name=field['name'], 
                        label=field['label'],
                        is_global=is_global  # âœ… NEW
                    )
                    
        except Exception as e:
            self._log_status(f"  ERROR in _get_picklist_fields: {str(e)}")
            
        return fields_dict
    
    def _is_global_picklist(self, field: dict) -> bool:
        """
        âœ… NEW METHOD: Detect if a picklist field uses a Global Value Set
        
        Args:
            field: Field describe dictionary
            
        Returns:
            True if field uses global value set, False otherwise
        """
        try:
            # Check for valueSet with restricted=true and valueSetName
            value_set = field.get('valueSet')
            
            if not value_set:
                return False
            
            # Global picklists have:
            # 1. restricted = true (or false, but usually true)
            # 2. valueSetName is not None (references global value set)
            
            value_set_name = value_set.get('valueSetName')
            
            # If valueSetName exists, it's a global picklist
            if value_set_name:
                return True
            
            return False
            
        except Exception as e:
            self._log_status(f"    âš ï¸  Error checking global picklist: {str(e)}")
            return False
    
    def _resolve_entity_definition_id(self, object_name: str) -> Optional[str]:
        """Resolve EntityDefinition ID for an object"""
        try:
            query = f"SELECT Id FROM EntityDefinition WHERE QualifiedApiName = '{object_name}'"
            url = f"{self.base_url}/services/data/v{self.api_version}/tooling/query/"
            response = requests.get(url, headers=self.headers, params={'q': query}, timeout=60)
            
            if response.status_code == 200:
                records = response.json().get('records', [])
                if records:
                    return records[0]['Id']
                    
        except Exception as e:
            self._log_status(f"  ERROR resolveEntityDefinitionId: {str(e)}")
            
        return None
    
    
    
    def _query_picklist_values_with_fallback(
        self, 
        object_name: str, 
        entity_def_id: Optional[str], 
        field_name: str
    ) -> List[PicklistValueDetail]:
        """Query picklist values using multiple fallback methods"""
        
        values = self._query_field_definition_tooling(object_name, field_name)
        if values:
            return values
        
        if entity_def_id:
            values = self._query_custom_field_tooling(entity_def_id, field_name)
            if values:
                return values
        
        values = self._query_custom_field_tooling_table_enum(object_name, field_name)
        if values:
            return values
        
        values = self._query_rest_describe_for_picklist(object_name, field_name)
        if values:
            return values
        
        return []
    
    def _query_field_definition_tooling(self, object_name: str, field_name: str) -> List[PicklistValueDetail]:
        """Query using FieldDefinition"""
        try:
            query = f"SELECT Metadata FROM FieldDefinition WHERE EntityDefinition.QualifiedApiName = '{object_name}' AND QualifiedApiName = '{field_name}'"
            url = f"{self.base_url}/services/data/v{self.api_version}/tooling/query/"
            response = requests.get(url, headers=self.headers, params={'q': query}, timeout=60)
            if response.status_code == 200:
                records = response.json().get('records', [])
                if records:
                    return self._parse_value_set(records[0].get('Metadata', {}))
        except Exception as e:
            self._log_status(f"      ERROR queryFieldDefinitionTooling: {str(e)}")
        return []
    
    def _query_custom_field_tooling(self, entity_def_id: str, field_name: str) -> List[PicklistValueDetail]:
        """Query using CustomField with entity ID"""
        try:
            dev_name = field_name[:-3] if field_name.endswith('__c') else field_name
            query = f"SELECT Metadata FROM CustomField WHERE TableEnumOrId = '{entity_def_id}' AND DeveloperName = '{dev_name}'"
            url = f"{self.base_url}/services/data/v{self.api_version}/tooling/query/"
            response = requests.get(url, headers=self.headers, params={'q': query}, timeout=60)
            if response.status_code == 200:
                records = response.json().get('records', [])
                if records:
                    return self._parse_value_set(records[0].get('Metadata', {}))
        except Exception as e:
            self._log_status(f"      ERROR queryCustomFieldTooling: {str(e)}")
        return []
    
    def _query_custom_field_tooling_table_enum(self, object_name: str, field_name: str) -> List[PicklistValueDetail]:
        """Query using CustomField with object name"""
        try:
            dev_name = field_name[:-3] if field_name.endswith('__c') else field_name
            query = f"SELECT Metadata FROM CustomField WHERE TableEnumOrId = '{object_name}' AND DeveloperName = '{dev_name}'"
            url = f"{self.base_url}/services/data/v{self.api_version}/tooling/query/"
            response = requests.get(url, headers=self.headers, params={'q': query}, timeout=60)
            if response.status_code == 200:
                records = response.json().get('records', [])
                if records:
                    return self._parse_value_set(records[0].get('Metadata', {}))
        except Exception as e:
            self._log_status(f"      ERROR queryCustomFieldToolingTableEnum: {str(e)}")
        return []
    
    def _query_rest_describe_for_picklist(self, object_name: str, field_name: str) -> List[PicklistValueDetail]:
        """Query using REST describe endpoint"""
        try:
            url = f"{self.base_url}/services/data/v{self.api_version}/sobjects/{object_name}/describe"
            response = requests.get(url, headers=self.headers, timeout=60)
            if response.status_code == 200:
                for field in response.json().get('fields', []):
                    if field['name'].lower() == field_name.lower():
                        return [
                            PicklistValueDetail(
                                label=pv.get('label', ''), 
                                value=pv.get('value', ''), 
                                is_active=pv.get('active', True)
                            ) for pv in field.get('picklistValues', [])
                        ]
        except Exception as e:
            self._log_status(f"      ERROR queryRestDescribeForPicklist: {str(e)}")
        return []
    
    def _parse_value_set(self, metadata: dict) -> List[PicklistValueDetail]:
        """Parse picklist values from metadata"""
        results = []
        try:
            value_set = metadata.get('valueSet', {})
            if not value_set:
                return results
            
            values = value_set.get('valueSetDefinition', {}).get('value', []) or value_set.get('value', [])
            for v in values:
                # is_active = bool(v.get('isActive', True))
                is_active_raw = v.get('isActive')
                if is_active_raw is None:
                    is_active = True  # Default for missing key
                else:
                    is_active = bool(is_active_raw)  # Convert any truthy/falsy value
                results.append(PicklistValueDetail(
                    label=v.get('label', ''), 
                    value=v.get('valueName') or v.get('value', ''), 
                    is_active=is_active
                ))
        except Exception as e:
            self._log_status(f"      ERROR parseValueSet: {str(e)}")
        return results
    
    def _create_excel_file(self, rows: List[List[str]], output_path: str) -> str:
        """
        âš ï¸ LEGACY: Create Excel file with formatted data (old method)
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Picklist Export"
        
        # âœ… Add headers as first row
        ws.append(self.PICKLIST_HEADERS)
        
        # Add data rows
        for row in rows:
            ws.append(row)
        
        # Format header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
        
        ws.freeze_panes = "A2"
        wb.save(output_path)
        
        self._log_status(f"âœ… Excel file created: {output_path}")
        self._log_status(f"âœ… Total data rows: {len(rows)}")
        return output_path
    
    def _log_status(self, message: str):
        """Log status message"""
        if self.sf_client.status_callback:
            self.sf_client.status_callback(message, verbose=True)
