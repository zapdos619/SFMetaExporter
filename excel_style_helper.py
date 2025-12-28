"""
Excel Styling Helper - Provides consistent formatting for Excel exports
UPDATED: Salesforce official blue colors + enhanced styling
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Tuple


class ExcelStyleHelper:
    """Helper class for consistent Excel styling across exports"""
    
    # ✅ NEW: Salesforce Official Brand Colors
    # Reference: https://www.salesforce.com/brand/color
    SALESFORCE_BLUE = "00A1E0"      # Salesforce Blue (lighter, brand primary)
    SALESFORCE_DARK_BLUE = "032D60" # Salesforce Dark Blue (darker navy)
    SALESFORCE_LIGHT_BLUE = "E6F4F9" # Very light blue for data rows (NEW)
    
    # Header colors (using official Salesforce blue)
    HEADER_BG_COLOR = "00A1E0"      # ✅ CHANGED: Lighter Salesforce blue
    HEADER_TEXT_COLOR = "FFFFFF"    # White text
    TITLE_BG_COLOR = "032D60"       # ✅ CHANGED: Darker Salesforce navy
    TITLE_TEXT_COLOR = "FFFFFF"     # White text
    
    # ✅ NEW: Alternate row coloring for better readability
    DATA_ROW_EVEN = "FFFFFF"        # White
    DATA_ROW_ODD = "F7FBFF"         # Very light blue tint
    
    @staticmethod
    def get_title_style():
        """Get style for title row (Row 1) - Darker Navy Blue"""
        return {
            'font': Font(
                name='Calibri',
                size=14,
                bold=True,
                color=ExcelStyleHelper.TITLE_TEXT_COLOR
            ),
            'fill': PatternFill(
                start_color=ExcelStyleHelper.TITLE_BG_COLOR,
                end_color=ExcelStyleHelper.TITLE_BG_COLOR,
                fill_type="solid"
            ),
            'alignment': Alignment(
                horizontal="left",
                vertical="center"
            ),
            'border': ExcelStyleHelper._get_border()
        }
    
    @staticmethod
    def get_info_style():
        """Get style for info row (Row 2) - Lighter Salesforce Blue"""
        return {
            'font': Font(
                name='Calibri',
                size=11,
                bold=False,
                color=ExcelStyleHelper.HEADER_TEXT_COLOR
            ),
            'fill': PatternFill(
                start_color=ExcelStyleHelper.HEADER_BG_COLOR,
                end_color=ExcelStyleHelper.HEADER_BG_COLOR,
                fill_type="solid"
            ),
            'alignment': Alignment(
                horizontal="left",
                vertical="center"
            ),
            'border': ExcelStyleHelper._get_border()
        }
    
    @staticmethod
    def get_header_style():
        """Get style for column headers (Row 3) - Lighter Salesforce Blue"""
        return {
            'font': Font(
                name='Calibri',
                size=12,
                bold=True,
                color=ExcelStyleHelper.HEADER_TEXT_COLOR
            ),
            'fill': PatternFill(
                start_color=ExcelStyleHelper.HEADER_BG_COLOR,
                end_color=ExcelStyleHelper.HEADER_BG_COLOR,
                fill_type="solid"
            ),
            'alignment': Alignment(
                horizontal="center",
                vertical="center"
            ),
            'border': ExcelStyleHelper._get_border()
        }
    
    @staticmethod
    def get_data_style(is_even_row=True):
        """
        Get style for data rows with optional alternating colors
        
        Args:
            is_even_row: If True, use even row color; if False, use odd row color
        """
        bg_color = ExcelStyleHelper.DATA_ROW_EVEN if is_even_row else ExcelStyleHelper.DATA_ROW_ODD
        
        return {
            'font': Font(
                name='Calibri',
                size=11
            ),
            'fill': PatternFill(
                start_color=bg_color,
                end_color=bg_color,
                fill_type="solid"
            ),
            'alignment': Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=True
            ),
            'border': ExcelStyleHelper._get_border(style='thin')
        }
    
    @staticmethod
    def _get_border(style='medium'):
        """Get border style"""
        side = Side(style=style, color="000000")
        return Border(left=side, right=side, top=side, bottom=side)
    
    @staticmethod
    def apply_style_to_cell(cell, style_dict):
        """Apply a style dictionary to a cell"""
        if 'font' in style_dict:
            cell.font = style_dict['font']
        if 'fill' in style_dict:
            cell.fill = style_dict['fill']
        if 'alignment' in style_dict:
            cell.alignment = style_dict['alignment']
        if 'border' in style_dict:
            cell.border = style_dict['border']
    
    @staticmethod
    def apply_style_to_row(ws, row_num, num_cols, style_dict):
        """Apply style to entire row"""
        for col_num in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col_num)
            ExcelStyleHelper.apply_style_to_cell(cell, style_dict)
    
    @staticmethod
    def add_title_row(ws, title: str, num_cols: int, row_num: int = 1):
        """
        Add formatted title row (Row 1) - Darker Navy Blue
        
        Args:
            ws: Worksheet
            title: Title text (e.g., "Salesforce Picklist Export")
            num_cols: Number of columns to merge
            row_num: Row number (default 1)
        """
        # Merge cells for title
        ws.merge_cells(start_row=row_num, start_column=1, 
                      end_row=row_num, end_column=num_cols)
        
        # Set title text
        cell = ws.cell(row=row_num, column=1)
        cell.value = title
        
        # Apply title style (dark navy blue)
        title_style = ExcelStyleHelper.get_title_style()
        ExcelStyleHelper.apply_style_to_cell(cell, title_style)
        
        # Apply style to merged cells (for borders)
        for col_num in range(2, num_cols + 1):
            cell = ws.cell(row=row_num, column=col_num)
            ExcelStyleHelper.apply_style_to_cell(cell, title_style)
        
        # Set row height
        ws.row_dimensions[row_num].height = 25
    
    @staticmethod
    def add_info_row(ws, object_label: str, object_api: str, 
                     record_count: int, num_cols: int, row_num: int = 2):
        """
        Add formatted info row (Row 2) - Lighter Salesforce Blue
        
        Args:
            ws: Worksheet
            object_label: Object label (e.g., "Account")
            object_api: Object API name (e.g., "Account")
            record_count: Number of records
            num_cols: Number of columns to merge
            row_num: Row number (default 2)
        """
        # Merge cells for info
        ws.merge_cells(start_row=row_num, start_column=1, 
                      end_row=row_num, end_column=num_cols)
        
        # Build info text
        export_date = datetime.now().strftime('%Y-%m-%d %H:%M')
        info_text = f"Object: {object_label} ({object_api}) | Total Records: {record_count} | Export Date: {export_date}"
        
        # Set info text
        cell = ws.cell(row=row_num, column=1)
        cell.value = info_text
        
        # Apply info style (lighter blue)
        info_style = ExcelStyleHelper.get_info_style()
        ExcelStyleHelper.apply_style_to_cell(cell, info_style)
        
        # Apply style to merged cells (for borders)
        for col_num in range(2, num_cols + 1):
            cell = ws.cell(row=row_num, column=col_num)
            ExcelStyleHelper.apply_style_to_cell(cell, info_style)
        
        # Set row height
        ws.row_dimensions[row_num].height = 20
    
    @staticmethod
    def add_header_row(ws, headers: List[str], row_num: int = 3):
        """
        Add formatted header row (Row 3) - Lighter Salesforce Blue
        
        Args:
            ws: Worksheet
            headers: List of column header names
            row_num: Row number (default 3)
        """
        header_style = ExcelStyleHelper.get_header_style()
        
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = header
            ExcelStyleHelper.apply_style_to_cell(cell, header_style)
        
        # Set row height
        ws.row_dimensions[row_num].height = 20
    
    @staticmethod
    def auto_adjust_column_widths(ws, headers: List[str], 
                                  max_width: int = 50, min_width: int = 10):
        """
        Auto-adjust column widths based on content
        
        Args:
            ws: Worksheet
            headers: List of column headers
            max_width: Maximum column width
            min_width: Minimum column width
        """
        for col_num, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_num)
            
            # Start with header length
            max_length = len(str(header))
            
            # Check data rows (sample first 100 for performance)
            for row_num in range(4, min(104, ws.max_row + 1)):
                cell = ws.cell(row=row_num, column=col_num)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            # Set adjusted width (with padding)
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[col_letter].width = adjusted_width
    
    @staticmethod
    def freeze_header_rows(ws, num_rows: int = 3):
        """
        Freeze top rows (title, info, headers)
        
        Args:
            ws: Worksheet
            num_rows: Number of rows to freeze (default 3)
        """
        ws.freeze_panes = ws.cell(row=num_rows + 1, column=1)
    
    @staticmethod
    def get_object_label(sf_client, object_api_name: str) -> str:
        """
        Get object label from Salesforce
        
        Args:
            sf_client: SalesforceClient instance
            object_api_name: Object API name
            
        Returns:
            Object label (fallback to API name if not found)
        """
        try:
            obj_describe = getattr(sf_client.sf, object_api_name).describe()
            return obj_describe.get('label', object_api_name)
        except:
            return object_api_name
    
    @staticmethod
    def sanitize_sheet_name(name: str, max_length: int = 31) -> str:
        r"""
        Sanitize sheet name for Excel compatibility
        
        Excel sheet names:
        - Max 31 characters
        - Cannot contain: : \ / ? * [ ]
        
        Args:
            name: Original name
            max_length: Maximum length (default 31)
            
        Returns:
            Sanitized sheet name
        """
        # Remove invalid characters
        invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
        sanitized = name
        for char in invalid_chars:
            sanitized = sanitized.replace(char, '_')
        
        # Truncate to max length
        if len(sanitized) > max_length:
            sanitized = sanitized[:max_length]
        
        # Ensure not empty
        if not sanitized:
            sanitized = "Sheet1"
        
        return sanitized